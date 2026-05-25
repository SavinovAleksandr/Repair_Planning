# -*- coding: utf-8 -*-
"""
Сборщик сводного графика ремонтов ЛЭП и сетевого оборудования.

Copyright (c) 2026 Савинов Александр, Сыктывкар. Все права защищены.

Принцип работы
--------------
Скрипт ищет в своей папке файлы:
    * Проект Арх РДУ.xlsx                                   — экспорт из ПК «Ремонты»
    * Проект Коми РДУ.xlsx                                  — экспорт из ПК «Ремонты»
    * Приоритет строк по группам для сводного графика.xlsx — справочник приоритетов

Если каких-то файлов в корне нет — пытается их найти в подпапке «Исходные материалы».

На выходе в той же корневой папке появляется файл:
    Сводный график ремонтов ЛЭП и сетевого оборудования на <месяц> <год> г.xlsx

Запуск
------
    python build_svod.py                    — собрать; год/месяц определяются автоматически
    python build_svod.py --year 2026        — указать год вручную
    python build_svod.py --no-normalize     — без текстовой нормализации
    python build_svod.py --collapse-preamble— дополнительно сворачивать преамбулы
                                              «Вывод в ремонт … для проведения …»
    python build_svod.py --dry-run          — ничего не сохранять, только отчёт
"""

from __future__ import annotations

import argparse
import difflib
import re
import shutil
import sys
from collections import Counter, defaultdict, OrderedDict
from copy import copy as _copy
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

# ------------------------------------------------------------------ КОНСТАНТЫ -

__copyright__ = "© Савинов Александр, Сыктывкар, 2026"

ROOT = Path(__file__).resolve().parent

FALLBACK_DIR = ROOT / "Исходные материалы"

FILE_ARKH = "Проект Арх РДУ.xlsx"
FILE_KOMI = "Проект Коми РДУ.xlsx"
FILE_PRIO = "Приоритет строк по группам для сводного графика.xlsx"

# Итоговое число колонок таблицы (A..Y).
TABLE_COLS = 25
LAST_COL_LETTER = get_column_letter(TABLE_COLS)

# Группы в порядке вывода.
GROUP_OGR = "OGR"          # Ограничения ОЗ
GROUP_LEP220 = "LEP220"    # ЛЭП 220 кВ
GROUP_PS220 = "PS220"      # ПС 220 кВ
GROUP_LEP110 = "LEP110"    # ЛЭП 110 кВ
GROUP_PS110 = "PS110"      # ПС 110 кВ
GROUP_ES = "ES"            # Электростанции
GROUP_ACHR = "ACHR"        # АЧР
GROUP_OTHER = "OTHER"      # Прочее (попадает всё, что не удалось классифицировать)

GROUP_ORDER = [
    GROUP_OGR,
    GROUP_LEP220,
    GROUP_PS220,
    GROUP_LEP110,
    GROUP_PS110,
    GROUP_ES,
    GROUP_ACHR,
    GROUP_OTHER,
]

GROUP_LABELS = {
    GROUP_OGR:    "Ограничения ОЗ",
    GROUP_LEP220: "ЛЭП 220 кВ",
    GROUP_PS220:  "ПС 220 кВ",
    GROUP_LEP110: "ЛЭП 110 кВ",
    GROUP_PS110:  "ПС 110 кВ",
    GROUP_ES:     "Электростанции",
    GROUP_ACHR:   "АЧР",
    GROUP_OTHER:  "Прочее (не классифицировано)",
}

RU_MONTHS_NOM = [
    "",  # dummy для 1-based
    "январь", "февраль", "март",     "апрель", "май",    "июнь",
    "июль",   "август",  "сентябрь", "октябрь", "ноябрь", "декабрь",
]

RU_MONTHS_SHORT = [
    "",
    "янв", "фев", "мар", "апр", "май", "июн",
    "июл", "авг", "сен", "окт", "ноя", "дек",
]

# ---- Высоты строк (пт) ----
ROW_HEIGHT_SECTION = 22.0       # заголовок группы (ЛЭП 220 кВ, …)
ROW_HEIGHT_SUBSECTION = 18.0    # подзаголовок объекта (ПС 220 кВ Вельск, …)
ROW_HEIGHT_TOC = 18.0           # строка оглавления

# ---- Стандартные объединения в строке оборудования ----------------------
# В проектах ПК «Ремонты» часть строк оборудования приходит без merge (или
# с нестандартной раскладкой). В своднике хочется единообразия: имя
# объекта занимает A:D, причины/условия — H:M, вид ремонта — N:O,
# ответственный — X:Y.
EQUIPMENT_MERGES: list[tuple[int, int]] = [
    (1, 4),    # A..D — наименование оборудования
    (8, 13),   # H..M — причины/условия
    (14, 15),  # N..O — вид ремонта / примечание
    (24, 25),  # X..Y — ответственный
]

# ---- Размеры шрифтов в строках данных (для расчёта высоты строки) -------
EQ_FONT_PT_A = 10.0   # колонка A (наименование)
EQ_FONT_PT_HN = 8.0   # колонки H и N (описания)

# ---- Заливка листа «Диаграмма» по виду ремонта (RGB без #) ----
GANTT_COLORS: dict[str, str] = {
    "ТР":  "B6D7A8",   # светло-зелёный
    "СР":  "FFE599",   # светло-жёлтый
    "КР":  "EA9999",   # розово-красный
    "ВПр": "9FC5E8",   # голубой
    "ИСП": "F9CB9C",   # оранжевый
    "ЗРР": "B4A7D6",   # сиреневый
    "БВР": "CCCCCC",   # серый
}
GANTT_COLOR_OTHER = "EEEEEE"
GANTT_COLOR_WEEKEND = "F2F2F2"
GANTT_SHEET_NAME = "Диаграмма"

DIFF_SHEET_NAME = "Сравнение с проектами"
DIFF_FILL_DELETED_ROW = "FFC7CE"   # светло-красная заливка удалённых строк
DIFF_FILL_NEW_ROW = "E2EFDA"       # светло-зелёная — новые строки в своднике
DIFF_FILL_DATE_CHG = "FFF2CC"      # жёлтая — изменённые даты
DIFF_COLOR_ADD = "008000"          # зелёный текст (добавления)
DIFF_COLOR_DEL = "FF0000"          # красный зачёркнутый (удаления)

BACKUP_DIR = ROOT / "backups"
SVOD_FILE_PREFIX = "Сводный график ремонтов"


@dataclass
class ProjectLayout:
    """Распознанная раскладка экспорта ПК «Ремонты» на листе проекта.

    Поля заполняются автоматически функцией `detect_project_layout` и
    позволяют читать проекты с разным числом строк шапки и смещёнными
    колонками дат / вида ремонта."""
    header_last: int = 6
    col_name: int = 1
    col_start: int = 6
    col_end: int = 7
    col_repair: int = 14
    table_cols: int = TABLE_COLS
    sheet_title: str = "Page1"


# ------------------------------------------------------------------ УТИЛИТЫ --

def find_file(name: str) -> Path:
    """Ищет файл в корне, затем в 'Исходные материалы'. Возвращает Path или None."""
    for base in (ROOT, FALLBACK_DIR):
        p = base / name
        if p.exists():
            return p
    return None


def parse_day_month(value, default_year: int) -> tuple[int, int, int] | None:
    """Разбирает '12.05.' / '12.05' / '12.05.2026' / datetime в (год, месяц, день).
    Возвращает None, если распарсить не удалось."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return (value.year, value.month, value.day)
    s = str(value).strip().rstrip(".").strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?$", s)
    if not m:
        return None
    day, mon = int(m.group(1)), int(m.group(2))
    year = int(m.group(3)) if m.group(3) else default_year
    if year < 100:
        year += 2000
    return (year, mon, day)


def _cell_text_with_merges(ws: Worksheet, row: int, col: int) -> str:
    """Возвращает текст ячейки; если ячейка внутри объединения и сама пустая —
    вернёт текст «владельца» объединения (top-left-ячейки)."""
    v = ws.cell(row, col).value
    if v is not None and str(v).strip() != "":
        return str(v)
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            owner = ws.cell(mr.min_row, mr.min_col).value
            if owner is not None:
                return str(owner)
            break
    return ""


def _hdr_text(ws: Worksheet, row: int, col: int) -> str:
    return re.sub(r"\s+", " ", _cell_text_with_merges(ws, row, col)).strip().lower()


def detect_project_layout(ws: Worksheet, filename: str = "") -> ProjectLayout:
    """Автоопределение раскладки экспорта ПК «Ремонты».

    Поддерживаются варианты с разным числом строк шапки (например 5–6 или
    3–4), другим именем листа и смещёнными колонками «Начало/Окончание» и
    «Вид ремонта» — главное, чтобы в шапке были узнаваемые подписи."""
    layout = ProjectLayout(sheet_title=ws.title)

    header_last: int | None = None
    start_cols: list[int] = []
    end_cols: list[int] = []

    scan_hi = min(25, ws.max_row + 1)
    for r in range(1, scan_hi):
        row_starts: list[int] = []
        row_ends: list[int] = []
        for c in range(1, TABLE_COLS + 1):
            t = _hdr_text(ws, r, c)
            if not t:
                continue
            if "начало" in t:
                row_starts.append(c)
            if "окончан" in t:
                row_ends.append(c)
        if row_starts and row_ends:
            header_last = r
            start_cols = row_starts
            end_cols = row_ends

    if header_last is None:
        # Запасной вариант: последняя строка блока с «наименован» + «разрешен».
        name_row = None
        for r in range(1, scan_hi):
            for c in range(1, TABLE_COLS + 1):
                if "наименован" in _hdr_text(ws, r, c):
                    name_row = r
                    break
            if name_row is not None:
                break
        header_last = (name_row + 1) if name_row else 6

    layout.header_last = header_last

    if start_cols:
        layout.col_start = start_cols[0]
        later_ends = [c for c in end_cols if c > layout.col_start]
        layout.col_end = later_ends[0] if later_ends else layout.col_start + 1

    name_found = False
    repair_found = False
    for r in range(max(1, header_last - 3), header_last + 1):
        for c in range(1, TABLE_COLS + 1):
            t = _hdr_text(ws, r, c)
            if not t:
                continue
            if not name_found and "наименован" in t:
                layout.col_name = c
                name_found = True
            if not repair_found and (
                    ("вид" in t and "ремонт" in t)
                    or ("аварийн" in t and "готов" in t)):
                layout.col_repair = c
                repair_found = True

    # Ширина таблицы — по самому широкому однострочному merge в данных.
    for r in range(header_last + 1, min(header_last + 80, ws.max_row + 1)):
        for mr in ws.merged_cells.ranges:
            if (mr.min_row == r and mr.max_row == r
                    and mr.min_col == 1 and mr.max_col >= layout.table_cols):
                layout.table_cols = max(layout.table_cols, mr.max_col)

    return layout


def layout_looks_valid(layout: ProjectLayout) -> bool:
    return (
        layout.header_last >= 2
        and layout.col_name >= 1
        and layout.col_start >= 1
        and layout.col_end > layout.col_start
    )


def validate_project_layout(layout: ProjectLayout, filename: str) -> None:
    """Проверяет, что раскладка похожа на экспорт ПК «Ремонты»."""
    errors: list[str] = []
    if not layout_looks_valid(layout):
        errors.append(
            "не удалось найти строку шапки с колонками «Начало» и «Окончание»"
        )
    if layout.col_name < 1:
        errors.append("не найдена колонка «Наименование …»")
    if layout.col_repair < 1:
        errors.append("не найдена колонка «Вид ремонта / аварийная готовность»")

    if errors:
        print()
        print(f"ОШИБКА: файл «{filename}» не похож на экспорт ПК «Ремонты».")
        for e in errors:
            print(f"  • {e}")
        print()
        print("Ожидается таблица с шапкой, где есть:")
        print("  • «Наименование …» (объект / оборудование);")
        print("  • «Начало (дата)» и «Окончание (дата)»;")
        print("  • «Вид ремонта» или «аварийная готовность».")
        print("Имя листа может отличаться от «Page1» — скрипт подберёт лист сам.")
        print("Если формат всё равно не распознан — приложите файл разработчику.")
        sys.exit(3)


def validate_project_template(ws: Worksheet, filename: str) -> ProjectLayout:
    """Проверяет лист и возвращает распознанную раскладку."""
    layout = detect_project_layout(ws, filename)
    validate_project_layout(layout, filename)
    return layout


def find_project_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """Выбирает лист с данными проекта: «Page1» или первый узнаваемый."""
    if "Page1" in wb.sheetnames:
        ws = wb["Page1"]
        if layout_looks_valid(detect_project_layout(ws)):
            return ws
    for name in wb.sheetnames:
        ws = wb[name]
        if layout_looks_valid(detect_project_layout(ws)):
            return ws
    return wb.active


def month_day_count(year: int, month: int) -> int:
    """Количество дней в указанном месяце."""
    if month == 12:
        nxt = datetime(year + 1, 1, 1)
    else:
        nxt = datetime(year, month + 1, 1)
    return (nxt - datetime(year, month, 1)).days


def copy_cell_style(src, dst):
    """Копирует стиль исходной ячейки в целевую (возможно, из другой книги)."""
    if src.has_style:
        dst.font = _copy(src.font)
        dst.fill = _copy(src.fill)
        dst.border = _copy(src.border)
        dst.alignment = _copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = _copy(src.protection)


def copy_cell(src, dst):
    dst.value = src.value
    copy_cell_style(src, dst)


def copy_row_full(src_ws: Worksheet, src_row: int,
                  dst_ws: Worksheet, dst_row: int,
                  ncols: int = TABLE_COLS):
    for c in range(1, ncols + 1):
        copy_cell(src_ws.cell(src_row, c), dst_ws.cell(dst_row, c))
    rh = src_ws.row_dimensions[src_row].height
    if rh is not None:
        dst_ws.row_dimensions[dst_row].height = rh


def copy_merges_in_row(src_ws: Worksheet, src_row: int,
                       dst_ws: Worksheet, dst_row: int,
                       ncols: int = TABLE_COLS):
    """Копирует объединения, находящиеся в указанной строке источника."""
    ranges = list(src_ws.merged_cells.ranges)
    for mr in ranges:
        if mr.min_row != src_row or mr.max_row != src_row:
            continue
        if mr.min_col > ncols:
            continue
        lo = mr.min_col
        hi = min(mr.max_col, ncols)
        rng = f"{get_column_letter(lo)}{dst_row}:{get_column_letter(hi)}{dst_row}"
        try:
            dst_ws.merge_cells(rng)
        except Exception:
            pass


def copy_column_widths(src_ws: Worksheet, dst_ws: Worksheet,
                       ncols: int = TABLE_COLS + 1):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        w = src_ws.column_dimensions[letter].width
        if w:
            dst_ws.column_dimensions[letter].width = w


# -------------------------------------------------------- ПАРСИНГ ПРОЕКТА ----

def is_section_row(ws: Worksheet, row: int,
                   ncols: int | None = None) -> bool:
    """Строка-подзаголовок: объединена на всю ширину (A..Y)."""
    if ncols is None:
        ncols = TABLE_COLS
    for mr in ws.merged_cells.ranges:
        if (mr.min_row == row and mr.max_row == row
                and mr.min_col == 1 and mr.max_col >= ncols):
            return True
    return False


def is_equipment_row(ws: Worksheet, row: int) -> bool:
    """Строка оборудования: в A что-то есть, при этом это не секция и не
    строка подписи (у подписей A пусто)."""
    a = ws.cell(row, 1).value
    if a is None or str(a).strip() == "":
        return False
    return not is_section_row(ws, row)


def find_data_bounds(ws: Worksheet, ncols: int = TABLE_COLS,
                     layout: ProjectLayout | None = None
                     ) -> tuple[int, int, int]:
    """Возвращает (header_last_row, data_last_row, signatures_start_row).

    Правила:
      * Шапка таблицы — до `layout.header_last` (по умолчанию 6).
      * Данные — строки с непустым A (секции или оборудование).
      * Подписи — начинаются после последней «data»-строки, могут содержать
        пустые промежутки между подписывающими лицами.
    """
    header_last = layout.header_last if layout else 6
    last_data_row = header_last
    for r in range(header_last + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is not None and str(a).strip() != "":
            last_data_row = r

    # sig_start — первая непустая строка после data-блока.
    sig_start = last_data_row + 1
    while sig_start <= ws.max_row:
        row_empty = True
        for c in range(1, ncols + 1):
            v = ws.cell(sig_start, c).value
            if v is not None and str(v).strip() != "":
                row_empty = False
                break
        if row_empty:
            sig_start += 1
        else:
            break

    data_last = last_data_row
    return header_last, data_last, sig_start


def extract_records(ws: Worksheet, rdu: str, default_year: int,
                    src_key: str,
                    layout: ProjectLayout | None = None) -> list[dict]:
    """Возвращает список записей с исходных строк оборудования.

    Каждая запись содержит ссылку на исходный лист и номер строки — это
    позволит затем скопировать её «как есть» (со всеми стилями и объединениями).
    """
    if layout is None:
        layout = detect_project_layout(ws)
    header_last, data_last, _sig_start = find_data_bounds(ws, layout=layout)
    recs: list[dict] = []
    current_section = None
    for r in range(header_last + 1, data_last + 1):
        a = ws.cell(r, layout.col_name).value
        if a is None or (isinstance(a, str) and a.strip() == ""):
            continue
        name = str(a).strip()
        if is_section_row(ws, r, layout.table_cols):
            current_section = name
            continue
        # строка оборудования
        start_raw = ws.cell(r, layout.col_start).value
        end_raw   = ws.cell(r, layout.col_end).value
        start = parse_day_month(start_raw, default_year)
        end   = parse_day_month(end_raw,   default_year)
        recs.append({
            "rdu":     rdu,                        # 'Арх' / 'Коми'
            "section": current_section or "",      # подзаголовок проекта
            "name":    name,                       # значение в столбце A
            "start":   start,                      # (y, m, d) или None
            "end":     end,                        # (y, m, d) или None
            "src_ws":  ws,
            "src_row": r,
            "src_key": src_key,                    # 'arkh' / 'komi' для отладки
            "layout":  layout,
        })
    return recs


# -------------------------------------------------- СПРАВОЧНИК ПРИОРИТЕТОВ ---

def load_priority(path: Path) -> dict:
    """Возвращает словарь с порядками объектов по группам.

    Правило разбора:
      * Заголовок раздела — строка, заканчивающаяся двоеточием
        (например «ПС 220 кВ:», «Электростанции:», «АЧР:»).
        Подсказки вида «сначала ПС 220 кВ ОЗ Архангельского РДУ:» тоже
        оканчиваются ":", но идентификатор раздела по ним не меняется.
      * Служебная строка «отсортировать даты начала…» — игнорируется.
      * Элементы вида «Ограничения ОЗ Архангельского РДУ» сами задают
        раздел OGR (в справочнике у этой группы нет отдельного заголовка).
      * Прочие строки — элементы списка текущего раздела.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    def section_of(low: str) -> str | None:
        if "ограничения оз" in low:            return "OGR"
        if low.startswith("лэп 220"):          return "LEP220"
        if low.startswith("пс 220"):           return "PS220"
        if low.startswith("лэп 110"):          return "LEP110"
        if low.startswith("пс 110"):           return "PS110"
        if low.startswith("электростанц"):     return "ES"
        if low.startswith("ачр"):              return "ACHR"
        return None

    current: str | None = None
    data: dict[str, list[str]] = defaultdict(list)

    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b is None:
            continue
        text = str(b).strip()
        if text == "" or text.startswith("Приоритет"):
            continue

        low = text.lower().rstrip(":").strip()

        if "отсортировать" in low:
            continue

        if text.rstrip().endswith(":"):
            # Заголовок раздела или внутренняя подсказка.
            sec = section_of(low)
            if sec:
                current = sec
            continue

        # Элементы списка «Ограничения ОЗ …» сами открывают раздел OGR
        # (отдельного заголовка группы в справочнике нет).
        if low.startswith("ограничения оз"):
            current = "OGR"
            data["OGR"].append(text)
            continue

        if current is not None:
            data[current].append(text)

    return {
        "OGR":    data.get("OGR", []),
        "LEP220": data.get("LEP220", []),
        "PS220":  data.get("PS220", []),
        "LEP110": data.get("LEP110", []),
        "PS110":  data.get("PS110", []),
        "ES":     data.get("ES", []),
        "ACHR":   data.get("ACHR", []),
    }


# ------------------------------------------------------------ КЛАССИФИКАЦИЯ --

RE_ACHR       = re.compile(r"(?i)(?:снижение объ[её]ма нагрузки|ачр)")
RE_OGRAN      = re.compile(r"(?i)ограничени\w*\s+режим")
RE_LINE       = re.compile(r"(?i)^\s*вл\s")
RE_220        = re.compile(r"(?i)220\s*кв")
RE_110        = re.compile(r"(?i)110\s*кв")
RE_PS_SECT    = re.compile(r"(?i)^\s*пс\s+(220|110)\s*кв")
RE_ES_SECT    = re.compile(r"(?i)(тэц|грэс)")  # 'ТЭЦ СЛПК', 'Сосногорская ТЭЦ', 'Печорская ГРЭС'

def classify(rec: dict) -> tuple[str, str]:
    """Возвращает (group_key, subgroup_label).
    subgroup_label — название ПС/Электростанции/ОЗ для групп, где это уместно,
    либо "" для «плоских» групп (ЛЭП, АЧР, Прочее)."""

    name = rec["name"] or ""
    section = rec["section"] or ""

    if RE_OGRAN.search(name):
        sub = f"Ограничения ОЗ {rec['rdu']} РДУ"
        return (GROUP_OGR, sub)

    if RE_ACHR.search(name):
        return (GROUP_ACHR, "")

    if RE_LINE.match(name):
        if RE_220.search(name):
            return (GROUP_LEP220, "")
        if RE_110.search(name):
            return (GROUP_LEP110, "")
        # ВЛ без явной отметки кВ — попробуем по секции
        if RE_220.search(section):
            return (GROUP_LEP220, "")
        if RE_110.search(section):
            return (GROUP_LEP110, "")
        return (GROUP_OTHER, "")

    # электростанция — определяем по секции
    if RE_ES_SECT.search(section):
        return (GROUP_ES, section.strip())

    m = RE_PS_SECT.match(section)
    if m:
        kv = m.group(1)
        if kv == "220":
            return (GROUP_PS220, section.strip())
        if kv == "110":
            return (GROUP_PS110, section.strip())

    return (GROUP_OTHER, section.strip())


# -------------------------------- ГРУППИРОВКА И СОРТИРОВКА РЕЗУЛЬТАТОВ ------

def _norm(s: str) -> str:
    """Нормализация названия объекта для сопоставления со справочником:
    удаляет лишние пробелы и кавычки-варианты, приводит к нижнему регистру."""
    s = s or ""
    s = s.replace("«", "").replace("»", "").replace('"', "").replace("'", "")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def subgroup_index(priority_list: list[str], label: str) -> int:
    """Возвращает индекс позиции объекта в справочнике, либо большое число,
    если объект не найден (такие уходят в конец группы)."""
    key = _norm(label)
    for i, item in enumerate(priority_list):
        if _norm(item) == key:
            return i
    return 10_000  # не найдено — в конец


def start_sort_key(rec: dict) -> tuple:
    s = rec.get("start")
    if s is None:
        # записи без даты — в самый конец своей группы
        return (9999, 99, 99)
    return s


def group_and_sort(records: list[dict], priority: dict) -> dict:
    """Возвращает OrderedDict: group_key -> list[record] (уже в порядке вывода).

    Для групп с подгруппами (ПС/Электростанции/Ограничения) записи внутри
    одной подгруппы идут подряд; порядок подгрупп задаётся справочником."""
    buckets = defaultdict(list)
    for rec in records:
        g, sub = classify(rec)
        rec["group"] = g
        rec["subgroup"] = sub
        buckets[g].append(rec)

    ordered: "OrderedDict[str, list[dict]]" = OrderedDict()
    unknown_warnings: list[str] = []

    for g in GROUP_ORDER:
        if g not in buckets or not buckets[g]:
            continue

        items = buckets[g]

        if g in (GROUP_LEP220, GROUP_LEP110, GROUP_ACHR, GROUP_OTHER):
            items.sort(key=start_sort_key)
        elif g == GROUP_OGR:
            # Сначала Арх, потом Коми; внутри — по дате.
            def ogr_key(r):
                rdu_order = 0 if r["rdu"] == "Арх" else 1
                return (rdu_order, ) + tuple(start_sort_key(r))
            items.sort(key=ogr_key)
        else:
            # PS220 / PS110 / ES — по справочнику, внутри подгруппы — по дате
            plist = priority.get(g, [])
            def sort_key(r):
                idx = subgroup_index(plist, r["subgroup"])
                return (idx, ) + tuple(start_sort_key(r))
            items.sort(key=sort_key)

            for r in items:
                if subgroup_index(plist, r["subgroup"]) >= 10_000:
                    msg = f"  [!] объект «{r['subgroup']}» (группа {GROUP_LABELS[g]}) не найден в справочнике приоритетов"
                    if msg not in unknown_warnings:
                        unknown_warnings.append(msg)

        ordered[g] = items

    if unknown_warnings:
        print("Предупреждения о неизвестных объектах:")
        for m in unknown_warnings:
            print(m)

    return ordered


# ------------------------------------------------------ СБОРКА ВЫХОДНОГО XLSX

def pick_month_year(records: list[dict], override_year: int | None) -> tuple[int, int]:
    """Определяет доминирующий месяц в заявках; год — по аргументу или по текущему."""
    months = Counter()
    years = Counter()
    for r in records:
        if r["start"]:
            y, m, _ = r["start"]
            months[m] += 1
            years[y] += 1
        elif r["end"]:
            y, m, _ = r["end"]
            months[m] += 1
            years[y] += 1
    month = months.most_common(1)[0][0] if months else datetime.now().month
    if override_year:
        year = override_year
    elif years:
        year = years.most_common(1)[0][0]
    else:
        year = datetime.now().year
    return month, year


def infer_schedule_from_filename(path: Path) -> tuple[int | None, int | None]:
    """Из имени «… на май 2026 г.xlsx» извлекает (month, year) или (None, None)."""
    name = path.stem.lower()
    year: int | None = None
    m_year = re.search(r"(?<![0-9])(\d{4})(?![0-9])", name)
    if m_year:
        year = int(m_year.group(1))
    month: int | None = None
    for i, mn in enumerate(RU_MONTHS_NOM):
        if i == 0:
            continue
        if mn in name:
            month = i
            break
    return month, year


def default_year_for_svod(svod_path: Path, year_hint: int | None = None) -> int:
    """Год для разбора дат вида «12.05.» в своднике."""
    _fn_month, fn_year = infer_schedule_from_filename(svod_path)
    return fn_year or year_hint or datetime.now().year


def find_style_rows(ws_komi: Worksheet,
                    layout: ProjectLayout | None = None) -> dict:
    """Находит в проекте Коми РДУ подходящие строки-образцы для стилей."""
    if layout is None:
        layout = detect_project_layout(ws_komi)
    header_last, data_last, sig_start = find_data_bounds(ws_komi, layout=layout)
    section_style_row = None
    equipment_style_row = None
    for r in range(header_last + 1, data_last + 1):
        if section_style_row is None and is_section_row(ws_komi, r, layout.table_cols):
            section_style_row = r
        if equipment_style_row is None and is_equipment_row(ws_komi, r):
            equipment_style_row = r
        if section_style_row and equipment_style_row:
            break
    return {
        "header_last": header_last,
        "data_last": data_last,
        "sig_start": sig_start,
        "section_style_row": section_style_row,
        "equipment_style_row": equipment_style_row,
        "layout": layout,
    }


def write_header(ws_komi: Worksheet, out_ws: Worksheet, header_last: int):
    """Копирует шапку (строки 1..header_last) из Коми проекта в выходной лист."""
    for r in range(1, header_last + 1):
        copy_row_full(ws_komi, r, out_ws, r)
    # Объединения в пределах шапки.
    for mr in ws_komi.merged_cells.ranges:
        if mr.min_row <= header_last and mr.max_row <= header_last:
            rng = f"{get_column_letter(mr.min_col)}{mr.min_row}:{get_column_letter(min(mr.max_col, TABLE_COLS + 1))}{mr.max_row}"
            try:
                out_ws.merge_cells(rng)
            except Exception:
                pass


def write_title(out_ws: Worksheet, month: int, year: int):
    """Обновляет тексты в шапке сводного графика.

    Конкретно:
      * Заголовок «Сводный график ремонта …» — в объединённой ячейке на
        3-й строке (обычно C3:X3).
      * Год в грифе «Утверждаю …» (D1) — заменяем «YYYY года» на текущий.
    """
    title_done = False
    for r in range(1, 7):
        for c in range(1, TABLE_COLS + 1):
            v = out_ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            if not title_done and v.strip().startswith("Сводный график"):
                new = (
                    "Сводный график ремонта ЛЭП и сетевого оборудования "
                    "операционной зоны Коми РДУ "
                    f"на {RU_MONTHS_NOM[month]} {year} г."
                )
                out_ws.cell(r, c).value = new
                title_done = True
                continue
            # Год в блоке «Утверждаю …» (обычно D1). Ищем шаблон «NNNN года»
            # в любом тексте шапки — и ставим актуальный год. Использованы
            # look-around вместо \b, т. к. перед годом может стоять «_»
            # (подчёркивания строки подписи), а \b после «_» не срабатывает.
            if re.search(r"(?<![0-9])\d{4}(?![0-9])\s*года", v):
                new = re.sub(r"(?<![0-9])\d{4}(?![0-9])(\s*года)",
                             rf"{year}\1", v)
                if new != v:
                    out_ws.cell(r, c).value = new


def _apply_section_vertical_center(ws: Worksheet, row: int) -> None:
    """Выравнивает текст заголовка/подзаголовка по центру строки."""
    cell = ws.cell(row, 1)
    al = cell.alignment
    cell.alignment = Alignment(
        horizontal=al.horizontal or "left",
        vertical="center",
        text_rotation=al.text_rotation,
        wrap_text=al.wrap_text,
        shrink_to_fit=al.shrink_to_fit,
        indent=al.indent,
    )


def write_style_row(out_ws: Worksheet, row: int, text: str,
                    src_ws: Worksheet, style_row: int,
                    height: float | None = None,
                    vertical_center: bool = True):
    """Пишет строку-заголовок/подзаголовок на всю ширину таблицы, копируя
    стиль из строки-образца проекта. Если указан `height` — принудительно
    выставляет высоту строки (pt); иначе копирует высоту из образца.

    `vertical_center=True` — текст по центру строки (как в ручном своднике
    для «ПС 220 кВ Микунь» и аналогичных подзаголовков)."""
    for c in range(1, TABLE_COLS + 1):
        copy_cell_style(src_ws.cell(style_row, c), out_ws.cell(row, c))
    cell = out_ws.cell(row, 1)
    cell.value = text
    if vertical_center:
        _apply_section_vertical_center(out_ws, row)
    rng = f"A{row}:{LAST_COL_LETTER}{row}"
    try:
        out_ws.merge_cells(rng)
    except Exception:
        pass
    if height is not None:
        out_ws.row_dimensions[row].height = height
    else:
        rh = src_ws.row_dimensions[style_row].height
        if rh is not None:
            out_ws.row_dimensions[row].height = rh


# ---------------------------------------------------------------------------
# Текстовая нормализация полей H (причины/условия) и N (вид ремонта, АГ, ...)
# ---------------------------------------------------------------------------
#
# Правила сформированы на основе сопоставления ручного сводного графика мая
# 2026 г. с исходными проектами. Каждое правило декларативно — регулярка +
# человекочитаемое имя (для отчёта) + действие.
#
# Действия:
#   * «H → N»  — короткая пометка в H вырезается и дописывается в конец N.
#   * «H drop» — короткая пометка в H вырезается без переноса.
#   * «ночь»   — «с включением на ночь» / «без включения на ночь» всегда
#                убирается из H и (если ещё нет) дописывается к N.
#   * «simple» — подстановки, применяемые и к H, и к N (общие нормализации).
#   * «преамбула» — опциональное сворачивание «Вывод в ремонт … для проведения
#                   <род. падеж> Y» → «<именительный падеж> Y» (флаг
#                   --collapse-preamble).

@dataclass
class NormOptions:
    """Настройки текстовой нормализации."""
    enabled: bool = True
    collapse_preamble: bool = False
    dry_run: bool = False


@dataclass
class NormStats:
    """Счётчики и детальный лог изменений для отчёта."""
    counts: Counter = field(default_factory=Counter)
    changes: list = field(default_factory=list)


# --- (1) Фразы из H, которые переносятся в конец N --------------------------

H_MOVE_TO_N_RULES: list[tuple[str, re.Pattern]] = [
    ("H→N «с переводом на ОШВ»", re.compile(r"с\s+переводом\s+на\s+ОШВ",
                                            re.IGNORECASE | re.UNICODE)),
    ("H→N «с переводом на ОВ»",  re.compile(r"с\s+переводом\s+на\s+ОВ",
                                            re.IGNORECASE | re.UNICODE)),
    ("H→N «Совместно с …»",       re.compile(r"Совместно\s+с\s+.+",
                                             re.IGNORECASE | re.UNICODE | re.DOTALL)),
]


# --- (2) Короткие «мусорные» ремарки, которые просто удаляются из H ---------

H_DROP_RULES: list[tuple[str, re.Pattern]] = [
    ("H убрано «не в транзите»",
     re.compile(r"не\s+в\s+транзите", re.IGNORECASE | re.UNICODE)),
    ("H убрано «с отключением без разбоки разъединителями»",
     re.compile(r"с\s+отключением\s+без\s+разб[оё]ки\s+разъединителями",
                re.IGNORECASE | re.UNICODE)),
]


# --- (3) Ночной режим — всегда переезжает из H в N --------------------------

NIGHT_RULES: list[tuple[str, re.Pattern]] = [
    ("«с включением на ночь» → N",
     re.compile(r"с\s+включением\s+на\s+ночь",  re.IGNORECASE | re.UNICODE)),
    ("«без включения на ночь» → N",
     re.compile(r"без\s+включения\s+на\s+ночь", re.IGNORECASE | re.UNICODE)),
]


# --- (4) Общие подстановки (применяются к H и N) ---------------------------

SIMPLE_SUBS: list[tuple[str, re.Pattern, object]] = [
    ("ТДТ → точки деления транзита",
     re.compile(r"\bТДТ\b", re.UNICODE), "точки деления транзита"),
    ("«А.Г.: ВЗ» → «А.Г.: ВЗ.»",
     re.compile(r"А\.Г\.:\s*ВЗ(?=\s)", re.UNICODE), "А.Г.: ВЗ."),
    ("«Включить» → «Включение»",
     re.compile(r"\bВключить\b", re.UNICODE), "Включение"),
    ("«Вывести в ремонт» → «Вывод в ремонт»",
     re.compile(r"\bВывести\s+в\s+ремонт\b", re.UNICODE), "Вывод в ремонт"),
    ("«NNNNг» → «NNNN г.»",
     re.compile(r"(\d{4})\s*г(?![а-яА-Я\.])", re.UNICODE), r"\1 г."),
    # Длинное тире между словами/числами (с пробелами вокруг).
    # Не трогает составные обозначения вида «АТ-2», «ВЛ-125», «28-30.05».
    ("Дефис между словами → длинное тире",
     re.compile(r"(?<=[A-Za-zА-Яа-я0-9])\s-\s(?=[A-Za-zА-Яа-я0-9])", re.UNICODE),
     " – "),
    # Пробел между числом и «кВ»: «110кВ» → «110 кВ».
    ("«NкВ» → «N кВ»",
     re.compile(r"(\d+)кВ\b", re.UNICODE), r"\1 кВ"),
    # Пробел между числом и «ч.»: «2ч» / «2ч.» → «2 ч.».
    # Не трогаем «часть», «часа», «часах» (следующая буква — русская).
    ("«Nч» → «N ч.»",
     re.compile(r"(?<![а-яА-Я\d])(\d+)\s*ч\.?(?![а-яА-Яa-zA-Z])", re.UNICODE),
     r"\1 ч."),
    # «ч. 30 м» / «ч.30 м» / «ч. 30м» → «ч. 30 мин.» — строго в контексте времени.
    ("«ч. Nм» → «ч. N мин.»",
     re.compile(r"(ч\.?)\s*(\d+)\s*м\.?(?![а-яА-Я])", re.UNICODE),
     r"\1 \2 мин."),
]


# --- (5) Опциональный коллапс преамбул в N ---------------------------------

PREAMBLE_RE = re.compile(
    r"(?P<prefix>.*?)"
    r"Вывод\w*\s+в\s+ремонт\s+"
    r"(?P<obj>.+?)"
    r"\s+(?P<link>на\s+время\s+проведения(?:\s+работ\s+по)?|"
    r"для\s+проведения|для|на\s+время)\s+"
    r"(?P<rest>.+)",
    re.IGNORECASE | re.UNICODE | re.DOTALL,
)

GEN_TO_NOM: dict[str, tuple[str, str]] = {
    "текущего ремонта":                 ("Текущий ремонт",       "текущий ремонт"),
    "среднего ремонта":                 ("Средний ремонт",       "средний ремонт"),
    "капитального ремонта":             ("Капитальный ремонт",   "капитальный ремонт"),
    "технического обслуживания":        ("Техническое обслуживание", "техническое обслуживание"),
    "профилактического восстановления": ("Профилактическое восстановление", "профилактическое восстановление"),
    "профилактическому восстановлению": ("Профилактическое восстановление", "профилактическое восстановление"),
    "испытаний":                        ("Проведение испытаний", "проведение испытаний"),
}


def _apply_h_rules(h: str, stats: NormStats) -> tuple[str, list[str]]:
    """Вычёркивает из H все предусмотренные фрагменты.
    Возвращает обновлённый H и список фрагментов для дописывания в N."""
    moves: list[str] = []
    text = h or ""

    # Ночь — всегда переносим
    for label, rx in NIGHT_RULES:
        m = rx.search(text)
        while m:
            moves.append(m.group(0))
            stats.counts[label] += 1
            text = text[:m.start()] + text[m.end():]
            m = rx.search(text)

    # Короткие хвосты H → N
    for label, rx in H_MOVE_TO_N_RULES:
        m = rx.search(text)
        if m:
            moves.append(m.group(0))
            stats.counts[label] += 1
            text = text[:m.start()] + text[m.end():]

    # Удаляемые пометки
    for label, rx in H_DROP_RULES:
        m = rx.search(text)
        if m:
            stats.counts[label] += 1
            text = text[:m.start()] + text[m.end():]

    # Чистка хвостов / повторов пробелов, но БЕЗ схлопывания пустых строк
    # между абзацами (пользователь мог поставить их осознанно).
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.strip()
    if re.fullmatch(r"[\s\.,;:]*", text or ""):
        text = ""

    return text, moves


def _append_moves_to_note(n: str, moves: list[str]) -> str:
    """Дописывает в конец N перенесённые из H фрагменты (без дублей)."""
    if not moves:
        return n or ""
    result = (n or "").rstrip()
    lowered = result.lower()
    for frag in moves:
        frag_norm = re.sub(r"\s+", " ", frag).strip()
        if not frag_norm:
            continue
        if frag_norm.lower() in lowered:
            continue
        if not result:
            result = frag_norm
        elif result.endswith(".") or result.endswith(":"):
            result = result + " " + frag_norm
        else:
            result = result + ". " + frag_norm
        lowered = result.lower()
    return result


def _apply_simple_subs(s: str, stats: NormStats) -> str:
    """Применяет список SIMPLE_SUBS к строке. Учитывает в статистике только
    фактические изменения (регекс может матчиться и на уже корректном тексте —
    такие «тождественные» срабатывания не считаем)."""
    if not s:
        return s
    for label, rx, repl in SIMPLE_SUBS:
        # Считаем, сколько матчей действительно меняют текст.
        if isinstance(repl, str):
            n_changed = sum(
                1 for m in rx.finditer(s)
                if m.expand(repl) != m.group(0)
            )
        else:
            n_changed = sum(
                1 for m in rx.finditer(s)
                if repl(m) != m.group(0)
            )
        if n_changed == 0:
            continue
        s = rx.sub(repl, s)
        stats.counts[label] += n_changed
    return s


def _collapse_preamble(n: str, stats: NormStats) -> str:
    """Сворачивает «Вывод в ремонт … для проведения <род> Y» в «<имен.> Y»."""
    if not n:
        return n
    m = PREAMBLE_RE.match(n)
    if not m:
        return n
    rest = m.group("rest")
    leading_key = None
    for k in sorted(GEN_TO_NOM.keys(), key=len, reverse=True):
        if re.match(r"^" + re.escape(k) + r"\b", rest, re.IGNORECASE):
            leading_key = k
            break
    if leading_key is None:
        return n

    stats.counts["Свёрнута преамбула «Вывод в ремонт … для проведения …»"] += 1
    cap_form, _ = GEN_TO_NOM[leading_key]
    rest_after = rest[len(leading_key):]
    # Остальные совпадения в rest_after приводим к lowercase-форме.
    for k in sorted(GEN_TO_NOM.keys(), key=len, reverse=True):
        lc_form = GEN_TO_NOM[k][1]
        rest_after = re.sub(r"\b" + re.escape(k) + r"\b", lc_form,
                            rest_after, flags=re.IGNORECASE)
    result = m.group("prefix") + cap_form + rest_after
    result = re.sub(r"[ \t]+", " ", result).rstrip()
    if not result.endswith("."):
        result += "."
    return result


def normalize_cells(h: str, n: str, opts: NormOptions, stats: NormStats,
                    row_label: str) -> tuple[str, str]:
    """Главная функция нормализации. Возвращает (new_H, new_N).

    Если opts.enabled = False — возвращает исходные значения без изменений.
    Все сработавшие правила учитываются в stats.counts; при любом изменении
    строка добавляется в stats.changes (для отчёта --dry-run)."""
    if not opts.enabled:
        return h or "", n or ""

    orig_h, orig_n = h or "", n or ""

    new_h, moves = _apply_h_rules(orig_h, stats)
    new_n = _append_moves_to_note(orig_n, moves)

    new_h = _apply_simple_subs(new_h, stats)
    new_n = _apply_simple_subs(new_n, stats)

    if opts.collapse_preamble:
        new_n = _collapse_preamble(new_n, stats)

    # Не логируем «пустое ≈ пустое» как изменение.
    def _changed(a: str, b: str) -> bool:
        if a == b:
            return False
        if (a or "").strip() == "" and (b or "").strip() == "":
            return False
        return True

    if _changed(orig_h, new_h) or _changed(orig_n, new_n):
        stats.changes.append({
            "row_label": row_label,
            "h_before":  orig_h,  "h_after": new_h,
            "n_before":  orig_n,  "n_after": new_n,
        })

    return new_h, new_n


def _sum_col_width(ws: Worksheet, lo: int, hi: int,
                   default: float = 8.43) -> float:
    """Суммарная ширина колонок [lo..hi] на листе в единицах Excel
    («количество символов шрифта по умолчанию»). Неустановленные ширины
    заменяются на `default` (стандарт Excel)."""
    total = 0.0
    for c in range(lo, hi + 1):
        w = ws.column_dimensions[get_column_letter(c)].width
        total += float(w) if w else default
    return total


def ensure_equipment_merges(ws: Worksheet, row: int,
                            merges: list[tuple[int, int]] = EQUIPMENT_MERGES
                            ) -> None:
    """Гарантирует, что в строке оборудования есть стандартные объединения
    A:D / H:M / N:O / X:Y. Если какой-то merge отсутствует — создаётся;
    если в диапазон «влез» меньший merge из источника — он снимается и
    переопределяется полностью."""
    # Существующие одностроковые объединения в этой строке.
    existing: list[tuple[int, int]] = [
        (mr.min_col, mr.max_col)
        for mr in ws.merged_cells.ranges
        if mr.min_row == row and mr.max_row == row
    ]
    for lo, hi in merges:
        if (lo, hi) in existing:
            continue
        # Снимаем частично пересекающиеся объединения внутри диапазона.
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row != row or mr.max_row != row:
                continue
            # пересечение по колонкам
            if not (mr.max_col < lo or mr.min_col > hi):
                try:
                    ws.unmerge_cells(str(mr))
                except Exception:
                    pass
        rng = f"{get_column_letter(lo)}{row}:{get_column_letter(hi)}{row}"
        try:
            ws.merge_cells(rng)
        except Exception:
            pass


def _count_wrapped_lines(text: str, chars_per_line: int) -> int:
    """Оценка числа визуальных строк, которые займёт `text` при wrap_text,
    если в одну строку помещается `chars_per_line` символов."""
    if not text:
        return 1
    s = str(text)
    chars_per_line = max(10, int(chars_per_line))
    total = 0
    for para in s.split("\n"):
        if not para:
            total += 1
            continue
        total += max(1, -(-len(para) // chars_per_line))
    return max(total, 1)


def estimate_eq_row_height(ws: Worksheet, a_text: str, h_text: str,
                           n_text: str) -> float:
    """Оценивает минимальную высоту строки оборудования (в пт), которой
    хватит, чтобы уместить любой из текстов A/H/N при `wrap_text=True`.

    Формула эмпирическая, но с запасом 10–15% — не вызывает наложений при
    открытии файла в Excel без ручной «Autofit Row Height»."""
    # Ширины merged-блоков в единицах Excel.
    w_a = _sum_col_width(ws, 1, 4)
    w_h = _sum_col_width(ws, 8, 13)
    w_n = _sum_col_width(ws, 14, 15)
    # Коэффициент: 1 unit ширины колонки ≈ 1 символ Calibri 11pt. Шрифт в
    # наших ячейках (Arial 8/10) уже, поэтому символов влезает больше.
    # 1.35 — консервативная оценка.
    k_hn = 1.35   # Arial 8
    k_a = 1.15    # Arial 10
    lines_a = _count_wrapped_lines(a_text, w_a * k_a)
    lines_h = _count_wrapped_lines(h_text, w_h * k_hn)
    lines_n = _count_wrapped_lines(n_text, w_n * k_hn)
    # Высота одной строки ≈ размер шрифта × 1.25–1.30.
    h_a = lines_a * EQ_FONT_PT_A * 1.30
    h_h = lines_h * EQ_FONT_PT_HN * 1.30
    h_n = lines_n * EQ_FONT_PT_HN * 1.30
    needed = max(h_a, h_h, h_n) + 3.0
    # Минимум — чтобы было не ниже одной строки; потолок — 409 pt
    # (стандартное ограничение Excel на высоту строки).
    return max(15.0, min(needed, 409.0))


def write_equipment_row(out_ws: Worksheet, dst_row: int, rec: dict,
                        opts: NormOptions, stats: NormStats,
                        force_height: bool = True):
    """Копирует строку оборудования из исходного листа, сохраняя стили и
    внутристрочные объединения (A:D для названия, N:O для примечания и т.п.),
    после чего нормализует текстовые поля H и N.

    Если `force_height=True` — высота строки рассчитывается исходя из длины
    текстов (чтобы в Excel не было визуальных наложений)."""
    src_ws = rec["src_ws"]
    src_row = rec["src_row"]
    copy_row_full(src_ws, src_row, out_ws, dst_row)
    copy_merges_in_row(src_ws, src_row, out_ws, dst_row)
    # Гарантируем стандартный набор объединений в строке данных.
    ensure_equipment_merges(out_ws, dst_row)

    h_cell = out_ws.cell(dst_row, 8)
    n_cell = out_ws.cell(dst_row, 14)
    row_label = f"R{dst_row} «{str(rec.get('name', '') or '')[:48]}»"
    new_h, new_n = normalize_cells(
        str(h_cell.value) if h_cell.value is not None else "",
        str(n_cell.value) if n_cell.value is not None else "",
        opts, stats, row_label,
    )
    if new_h != (h_cell.value or ""):
        h_cell.value = new_h if new_h else None
    if new_n != (n_cell.value or ""):
        n_cell.value = new_n if new_n else None

    # Гарантируем перенос текста в H и N (для корректного отображения
    # многострочных описаний).
    for cell in (h_cell, n_cell):
        al = cell.alignment
        if not al.wrap_text:
            cell.alignment = Alignment(
                horizontal=al.horizontal, vertical=al.vertical,
                text_rotation=al.text_rotation, wrap_text=True,
                shrink_to_fit=al.shrink_to_fit, indent=al.indent,
            )

    # Высота строки: берём максимум из скопированной и расчётной. Это
    # спасает от двух крайностей: (1) очень длинные тексты в Ограничениях
    # ОЗ (копия из исходника даёт 409 pt — сохраняем); (2) проекты, где
    # высота не выставлена или занижена — добираем расчётом.
    if force_height:
        a_text = str(out_ws.cell(dst_row, 1).value or "")
        est = estimate_eq_row_height(out_ws, a_text,
                                     str(new_h or ""), str(new_n or ""))
        existing = out_ws.row_dimensions[dst_row].height or 0.0
        out_ws.row_dimensions[dst_row].height = max(existing, est)
        try:
            out_ws.row_dimensions[dst_row].customHeight = True
        except Exception:
            pass


def write_signatures(ws_komi: Worksheet, out_ws: Worksheet,
                     sig_start: int, dst_start: int) -> int:
    """Переносит блок подписей из Коми РДУ после итоговых строк данных.
    Возвращает индекс строки после последнего перенесённого ряда."""
    sig_end = ws_komi.max_row
    for i, r in enumerate(range(sig_start, sig_end + 1)):
        dst_r = dst_start + i
        copy_row_full(ws_komi, r, out_ws, dst_r)
        copy_merges_in_row(ws_komi, r, out_ws, dst_r)
    return dst_start + (sig_end - sig_start + 1)


def write_toc(out_ws: Worksheet, toc_row: int,
              group_anchors: dict[str, int]) -> None:
    """Пишет в строке `toc_row` оглавление: по ячейке на каждую непустую
    группу с гиперссылкой на строку её заголовка."""
    if not group_anchors:
        return

    ordered = [g for g in GROUP_ORDER if g in group_anchors]
    n = len(ordered)
    if n == 0:
        return

    # Равномерно распределяем непустые группы по 25 колонкам.
    base_width = TABLE_COLS // n
    extra = TABLE_COLS - base_width * n
    spans: list[tuple[int, int]] = []
    col = 1
    for i in range(n):
        w = base_width + (1 if i < extra else 0)
        spans.append((col, col + w - 1))
        col += w

    link_font = Font(name="Calibri", size=11, bold=True, color="0563C1",
                     underline="single")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    out_ws.row_dimensions[toc_row].height = ROW_HEIGHT_TOC
    for (lo, hi), g in zip(spans, ordered):
        anchor = group_anchors[g]
        cell = out_ws.cell(toc_row, lo)
        cell.value = f"{GROUP_LABELS[g]} (стр. {anchor})"
        cell.font = link_font
        cell.alignment = center
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate,
            location=f"Page1!A{anchor}",
            display=cell.value,
        )
        if hi > lo:
            rng = (f"{get_column_letter(lo)}{toc_row}:"
                   f"{get_column_letter(hi)}{toc_row}")
            try:
                out_ws.merge_cells(rng)
            except Exception:
                pass


def _vid_remonta(n_text: str) -> str:
    """Извлекает короткий код вида ремонта из текста N (ВПр/ТР/СР/КР/ИСП/ЗРР/БВР).
    Возвращает пустую строку, если код не распознан."""
    if not n_text:
        return ""
    m = re.match(r"\s*(ВПр|ТР|СР|КР|ИСП|ЗРР|БВР)\b", n_text)
    return m.group(1) if m else ""


def _gantt_day_span(rec: dict, month: int, year: int,
                    scale_start: datetime, scale_end: datetime
                    ) -> tuple[int, int] | None:
    """По start/end записи возвращает (колонка_нач, колонка_кон) на шкале Ганта
    (индексы от 1) относительно scale_start. None — если дат нет или они
    полностью вне шкалы."""
    s, e = rec.get("start"), rec.get("end")
    if not s and not e:
        return None
    if s:
        sd = datetime(s[0], s[1], s[2])
    else:
        sd = datetime(year, month, 1)
    if e:
        ed = datetime(e[0], e[1], e[2])
    else:
        ed = sd
    if ed < sd:
        sd, ed = ed, sd
    if ed < scale_start or sd > scale_end:
        return None
    if sd < scale_start:
        sd = scale_start
    if ed > scale_end:
        ed = scale_end
    col_start = (sd - scale_start).days + 1
    col_end   = (ed - scale_start).days + 1
    return col_start, col_end


def build_gantt_sheet(out_wb: openpyxl.Workbook, gantt_items: list[dict],
                      month: int, year: int) -> None:
    """Добавляет в книгу лист «Диаграмма» с Гант-календарём.
    `gantt_items` — список словарей {row, group, rec}, в том же порядке, что
    записи на основном листе."""
    ws = out_wb.create_sheet(GANTT_SHEET_NAME)

    if not gantt_items:
        ws.cell(1, 1).value = "Нет строк для диаграммы."
        return

    # --- Шкала: от самой ранней start до самой поздней end, но гарантированно
    # включаем весь целевой месяц.
    month_start = datetime(year, month, 1)
    month_end = datetime(year, month, month_day_count(year, month))
    scale_start = month_start
    scale_end = month_end
    for it in gantt_items:
        s, e = it["rec"].get("start"), it["rec"].get("end")
        if s:
            d = datetime(s[0], s[1], s[2])
            if d < scale_start:
                scale_start = d
        if e:
            d = datetime(e[0], e[1], e[2])
            if d > scale_end:
                scale_end = d

    n_days = (scale_end - scale_start).days + 1

    COL_NAME = 1       # A — имя объекта
    COL_VID  = 2       # B — код вида ремонта
    COL_DAYS = 3       # C — первый день шкалы
    last_days_col = COL_DAYS + n_days - 1
    legend_col = last_days_col + 2   # пустая колонка-разрыв + легенда

    ws.column_dimensions[get_column_letter(COL_NAME)].width = 44
    ws.column_dimensions[get_column_letter(COL_VID)].width = 7
    for c in range(COL_DAYS, last_days_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 3.2

    thin_font = Font(name="Calibri", size=9)
    bold_font = Font(name="Calibri", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Строка 1: месяцы (объединённые) ---
    cur = scale_start.replace(day=1)
    while cur <= scale_end:
        if cur.month == 12:
            nxt = cur.replace(year=cur.year + 1, month=1)
        else:
            nxt = cur.replace(month=cur.month + 1)
        seg_start = max(cur, scale_start)
        seg_end = min(nxt - timedelta(days=1), scale_end)
        col_from = COL_DAYS + (seg_start - scale_start).days
        col_to = COL_DAYS + (seg_end - scale_start).days
        cell = ws.cell(1, col_from)
        cell.value = f"{RU_MONTHS_SHORT[cur.month]} {cur.year}"
        cell.font = bold_font
        cell.alignment = center
        if col_to > col_from:
            rng = (f"{get_column_letter(col_from)}1:"
                   f"{get_column_letter(col_to)}1")
            try:
                ws.merge_cells(rng)
            except Exception:
                pass
        cur = nxt

    # --- Строка 2: числа дней + Строка 3: день недели ---
    weekend_fill = PatternFill(start_color=GANTT_COLOR_WEEKEND,
                               end_color=GANTT_COLOR_WEEKEND,
                               fill_type="solid")
    wday_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for i in range(n_days):
        d = scale_start + timedelta(days=i)
        col = COL_DAYS + i
        c2 = ws.cell(2, col)
        c2.value = d.day
        c2.font = thin_font
        c2.alignment = center
        c3 = ws.cell(3, col)
        c3.value = wday_names[d.weekday()]
        c3.font = thin_font
        c3.alignment = center
        if d.weekday() >= 5:
            c2.fill = weekend_fill
            c3.fill = weekend_fill

    # Заголовки A/B.
    ws.cell(1, COL_NAME).value = "Объект"
    ws.cell(1, COL_NAME).font = bold_font
    ws.cell(1, COL_NAME).alignment = center
    ws.cell(1, COL_VID).value = "Вид"
    ws.cell(1, COL_VID).font = bold_font
    ws.cell(1, COL_VID).alignment = center
    try:
        ws.merge_cells(f"{get_column_letter(COL_NAME)}1:"
                       f"{get_column_letter(COL_NAME)}3")
        ws.merge_cells(f"{get_column_letter(COL_VID)}1:"
                       f"{get_column_letter(COL_VID)}3")
    except Exception:
        pass

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 14

    # --- Строки данных ---
    row = 4
    for it in gantt_items:
        rec = it["rec"]
        name = str(rec.get("name") or "").strip()
        sub = str(rec.get("subgroup") or "").strip()
        display = f"{sub}: {name}" if sub and sub.lower() != name.lower() else name

        n_text = ""
        try:
            src_ws = rec["src_ws"]
            src_row = rec["src_row"]
            col_n = rec.get("layout")
            if isinstance(col_n, ProjectLayout):
                col_n = col_n.col_repair
            else:
                col_n = detect_project_layout(src_ws).col_repair
            n_text = _cell_text_with_merges(src_ws, src_row, col_n)
        except Exception:
            pass
        vid = _vid_remonta(n_text)

        ws.cell(row, COL_NAME).value = display
        ws.cell(row, COL_NAME).font = thin_font
        ws.cell(row, COL_NAME).alignment = left
        ws.cell(row, COL_VID).value = vid or "—"
        ws.cell(row, COL_VID).font = thin_font
        ws.cell(row, COL_VID).alignment = center

        # Подкрашиваем выходные в строке данных тоже.
        for i in range(n_days):
            d = scale_start + timedelta(days=i)
            if d.weekday() >= 5:
                ws.cell(row, COL_DAYS + i).fill = weekend_fill

        span = _gantt_day_span(rec, month, year, scale_start, scale_end)
        if span:
            color = GANTT_COLORS.get(vid, GANTT_COLOR_OTHER)
            fill = PatternFill(start_color=color, end_color=color,
                               fill_type="solid")
            for c in range(COL_DAYS + span[0] - 1, COL_DAYS + span[1]):
                ws.cell(row, c).fill = fill
        row += 1

    # --- Легенда ---
    ws.cell(1, legend_col).value = "Легенда"
    ws.cell(1, legend_col).font = bold_font
    ws.cell(1, legend_col).alignment = center
    legend_rows = [
        ("ТР",  "Текущий ремонт"),
        ("СР",  "Средний ремонт"),
        ("КР",  "Капитальный ремонт"),
        ("ВПр", "Внеплановый ремонт"),
        ("ИСП", "Испытания"),
        ("ЗРР", "Заявка РР"),
        ("БВР", "Без вывода в ремонт"),
        ("—",   "Прочее / код не распознан"),
    ]
    for i, (code, desc) in enumerate(legend_rows):
        r = 2 + i
        color = GANTT_COLORS.get(code, GANTT_COLOR_OTHER)
        c1 = ws.cell(r, legend_col)
        c1.value = code
        c1.font = thin_font
        c1.alignment = center
        c1.fill = PatternFill(start_color=color, end_color=color,
                              fill_type="solid")
        c2 = ws.cell(r, legend_col + 1)
        c2.value = desc
        c2.font = thin_font
        c2.alignment = left
    ws.column_dimensions[get_column_letter(legend_col)].width = 6
    ws.column_dimensions[get_column_letter(legend_col + 1)].width = 28

    # Закрепление областей: под шапкой и справа от столбцов-идентификаторов.
    ws.freeze_panes = ws.cell(4, COL_DAYS).coordinate

    # Печать — альбомная, вписать в 1 страницу по ширине.
    try:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    ws.print_area = (f"A1:{get_column_letter(legend_col + 1)}{row - 1}")


def build_output(priority: dict, records: list[dict],
                 ws_komi: Worksheet, ws_arkh: Worksheet | None,
                 month: int, year: int,
                 opts: NormOptions, stats: NormStats,
                 apply_sort: bool = True,
                 apply_toc: bool = True,
                 apply_heights: bool = True,
                 apply_gantt: bool = True) -> openpyxl.Workbook:
    """Собирает итоговую книгу.

    Флаги-стадии позволяют выключить отдельные преобразования:
      * apply_sort    — группировать и сортировать по приоритетам. Если False —
                        строки идут в порядке `records`, но классификация (для
                        заголовков групп) выполняется всегда.
      * apply_toc     — писать строку-оглавление со ссылками на группы.
      * apply_heights — фиксировать высоту строк-заголовков (секций/подсекций).
      * apply_gantt   — создавать лист «Диаграмма».
    Нормализация текста управляется `opts.enabled`.
    """
    style_info = find_style_rows(ws_komi)

    out_wb = openpyxl.Workbook()
    # удаляем стандартный лист «Sheet» и создаём «Page1», чтобы совпадало с проектами
    out_wb.remove(out_wb.active)
    out_ws = out_wb.create_sheet("Page1")

    # Ширины колонок — как в проекте Коми РДУ.
    copy_column_widths(ws_komi, out_ws, ncols=26)
    # параметры страницы
    try:
        out_ws.page_setup = _copy(ws_komi.page_setup)
        out_ws.print_options = _copy(ws_komi.print_options)
        out_ws.page_margins = _copy(ws_komi.page_margins)
        out_ws.sheet_properties.pageSetUpPr = _copy(ws_komi.sheet_properties.pageSetUpPr)
    except Exception:
        pass

    # Шапка.
    write_header(ws_komi, out_ws, style_info["header_last"])
    write_title(out_ws, month, year)

    # Группированные записи.
    if apply_sort:
        grouped = group_and_sort(records, priority)
    else:
        # Только классификация, без переупорядочивания: сохраняем порядок
        # строк из `records`, но раскладываем по корзинам согласно classify().
        buckets: "OrderedDict[str, list[dict]]" = OrderedDict()
        for rec in records:
            g, sub = classify(rec)
            rec["group"] = g
            rec["subgroup"] = sub
            buckets.setdefault(g, []).append(rec)
        grouped = OrderedDict()
        for g in GROUP_ORDER:
            if g in buckets and buckets[g]:
                grouped[g] = buckets[g]

    sect_style_row = style_info["section_style_row"]
    # Резервируем строку под оглавление; фактический текст TOC запишем в конце,
    # когда будут известны позиции всех заголовков групп.
    toc_row = style_info["header_last"] + 1
    cur = toc_row + 1

    group_anchors: dict[str, int] = {}
    # Порядок записей в том же виде, в котором они идут на листе (для Гант-листа).
    gantt_items: list[dict] = []

    section_h = ROW_HEIGHT_SECTION if apply_heights else None
    subsection_h = ROW_HEIGHT_SUBSECTION if apply_heights else None

    for g in GROUP_ORDER:
        if g not in grouped or not grouped[g]:
            continue
        # Заголовок группы.
        group_anchors[g] = cur
        write_style_row(out_ws, cur, GROUP_LABELS[g], ws_komi, sect_style_row,
                        height=section_h)
        cur += 1

        items = grouped[g]

        if g in (GROUP_PS220, GROUP_PS110, GROUP_ES, GROUP_OGR):
            # Подзаголовки по объектам.
            current_sub = None
            for r in items:
                if r["subgroup"] != current_sub:
                    current_sub = r["subgroup"]
                    if current_sub:
                        write_style_row(out_ws, cur, current_sub, ws_komi,
                                        sect_style_row,
                                        height=subsection_h)
                        cur += 1
                write_equipment_row(out_ws, cur, r, opts, stats)
                gantt_items.append({"row": cur, "group": g, "rec": r})
                cur += 1
        else:
            # «Плоские» группы (ЛЭП, АЧР, Прочее).
            for r in items:
                write_equipment_row(out_ws, cur, r, opts, stats)
                gantt_items.append({"row": cur, "group": g, "rec": r})
                cur += 1

    # Оглавление (гиперссылки на строки заголовков групп).
    if apply_toc:
        write_toc(out_ws, toc_row, group_anchors)

    # Подписи.
    write_signatures(ws_komi, out_ws, style_info["sig_start"], cur)

    # установим область печати (A..Y)
    out_ws.print_area = f"A1:{LAST_COL_LETTER}{out_ws.max_row}"

    # Второй лист — Гант-календарь.
    if apply_gantt:
        build_gantt_sheet(out_wb, gantt_items, month, year)

    # Основной лист должен открываться первым.
    out_wb.active = 0

    return out_wb


# ---------------------------------------------------------------------------
# Работа с уже существующим сводником: парсер, резервные копии, inplace-стадии
# ---------------------------------------------------------------------------


def find_existing_svod(root: Path) -> Path | None:
    """Возвращает путь к последнему (по mtime) файлу «Сводный график …xlsx»
    в корне проекта. None — если файла нет."""
    candidates = [
        p for p in root.glob(f"{SVOD_FILE_PREFIX}*.xlsx")
        if not p.name.startswith("~$")  # временный lock-файл Excel
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def make_backup(path: Path, log=print) -> Path | None:
    """Кладёт копию `path` в `backups/<timestamp>__<имя>.xlsx`. Возвращает путь
    к копии. Если файла нет — возвращает None."""
    if not path.exists():
        return None
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dst = BACKUP_DIR / f"{ts}__{path.name}"
    shutil.copy2(path, dst)
    log(f"Резервная копия: backups/{dst.name}")
    return dst


def restore_latest_backup(svod_path: Path, log=print) -> Path | None:
    """Восстанавливает `svod_path` из последней подходящей копии в `backups/`.
    Возвращает путь к восстановленному файлу либо None, если копий нет."""
    if not BACKUP_DIR.exists():
        log("Папка backups/ не найдена — нечего восстанавливать.")
        return None
    prefix = svod_path.name
    candidates = sorted(
        BACKUP_DIR.glob(f"*__{prefix}"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        log(f"В backups/ нет копий «{prefix}».")
        return None
    latest = candidates[0]
    # Сначала — копия текущего файла (на случай неудачного отката).
    if svod_path.exists():
        make_backup(svod_path, log=log)
    shutil.copy2(latest, svod_path)
    log(f"Восстановлено из: backups/{latest.name}")
    return svod_path


def is_toc_row(ws: Worksheet, row: int) -> bool:
    """Строка-оглавление в сводника: хотя бы одна ячейка содержит гиперссылку
    с локацией «Page1!A…»."""
    for c in range(1, TABLE_COLS + 1):
        cell = ws.cell(row, c)
        hl = getattr(cell, "hyperlink", None)
        if hl is None:
            continue
        loc = getattr(hl, "location", None) or ""
        if "Page1" in str(loc):
            return True
    return False


def extract_records_from_svod(ws_svod: Worksheet, default_year: int,
                              src_key: str = "svod") -> list[dict]:
    """Аналог `extract_records`, но для уже сгенерированного сводника.

    Пропускает строку оглавления и строки-заголовки групп (таких подгрупп
    нет в проектах, они появляются только в своднике). Подзаголовки по
    объектам (ПС/Электростанции) становятся current_section — как и в
    исходных проектах."""
    header_last, data_last, _sig_start = find_data_bounds(ws_svod)
    layout = detect_project_layout(ws_svod)
    recs: list[dict] = []
    current_section = ""
    group_label_lc = {v.strip().lower() for v in GROUP_LABELS.values()}

    for r in range(header_last + 1, data_last + 1):
        if is_toc_row(ws_svod, r):
            continue
        a = ws_svod.cell(r, 1).value
        if a is None or (isinstance(a, str) and a.strip() == ""):
            continue
        name = str(a).strip()

        if is_section_row(ws_svod, r):
            if name.strip().lower() in group_label_lc:
                # Заголовок группы верхнего уровня — данных не содержит.
                current_section = ""
                continue
            current_section = name
            continue

        # Определяем, из какого РДУ запись. В своднике это признак косвенный:
        # секция-подзаголовок может содержать «Архангельского» / «Коми».
        rdu = "Коми"
        sec_lc = current_section.lower()
        if "арх" in sec_lc:
            rdu = "Арх"
        elif "коми" in sec_lc:
            rdu = "Коми"

        start_raw = _cell_text_with_merges(ws_svod, r, layout.col_start)
        end_raw = _cell_text_with_merges(ws_svod, r, layout.col_end)
        start = parse_day_month(start_raw, default_year)
        end = parse_day_month(end_raw, default_year)

        recs.append({
            "rdu": rdu,
            "section": current_section,
            "name": name,
            "start": start,
            "end": end,
            "src_ws": ws_svod,
            "src_row": r,
            "src_key": src_key,
            "layout": layout,
        })
    return recs


def _save_with_backup(wb: openpyxl.Workbook, out_path: Path, log=print):
    """Сохраняет книгу по пути `out_path`, предварительно забэкапив старый
    файл (если он был). Понятно отрабатывает PermissionError."""
    if out_path.exists():
        make_backup(out_path, log=log)
    try:
        wb.save(out_path)
    except PermissionError:
        raise RuntimeError(
            f"Не удаётся сохранить «{out_path.name}» — вероятно, файл открыт "
            f"в Excel. Закройте его и попробуйте ещё раз."
        )
    log(f"Сохранено: {out_path.name}")


# --- Стадии, выполняемые «поверх» уже существующего сводника -----------------

def stage_normalize_inplace(svod_path: Path, opts: NormOptions,
                            stats: NormStats, log=print) -> None:
    """Нормализует текст в столбцах H/N непосредственно в файле `svod_path`."""
    log(f"Нормализация текста: {svod_path.name}")
    wb = openpyxl.load_workbook(svod_path)
    if "Page1" not in wb.sheetnames:
        raise RuntimeError("В файле нет листа «Page1» — это не похоже на сводный график.")
    ws = wb["Page1"]
    header_last, data_last, _ = find_data_bounds(ws)
    opts.enabled = True  # стадия явно включает нормализацию
    for r in range(header_last + 1, data_last + 1):
        if is_toc_row(ws, r) or is_section_row(ws, r):
            continue
        if not is_equipment_row(ws, r):
            continue
        h_cell = ws.cell(r, 8)
        n_cell = ws.cell(r, 14)
        row_label = f"R{r} «{_short(str(ws.cell(r, 1).value or ''), 48)}»"
        new_h, new_n = normalize_cells(
            str(h_cell.value) if h_cell.value is not None else "",
            str(n_cell.value) if n_cell.value is not None else "",
            opts, stats, row_label,
        )
        if new_h != (h_cell.value or ""):
            h_cell.value = new_h if new_h else None
        if new_n != (n_cell.value or ""):
            n_cell.value = new_n if new_n else None
    _save_with_backup(wb, svod_path, log=log)


def stage_build_toc_inplace(svod_path: Path, log=print) -> None:
    """Пересоздаёт строку оглавления в уже существующем своднике."""
    log(f"Оглавление: {svod_path.name}")
    wb = openpyxl.load_workbook(svod_path)
    if "Page1" not in wb.sheetnames:
        raise RuntimeError("В файле нет листа «Page1».")
    ws = wb["Page1"]
    header_last, data_last, _ = find_data_bounds(ws)

    # Найдём якоря: строки-секции, текст которых совпадает с названием одной
    # из групп верхнего уровня.
    group_anchors: dict[str, int] = {}
    label_to_key = {v.strip().lower(): k for k, v in GROUP_LABELS.items()}
    for r in range(header_last + 1, data_last + 1):
        if not is_section_row(ws, r):
            continue
        text = str(ws.cell(r, 1).value or "").strip().lower()
        key = label_to_key.get(text)
        if key and key not in group_anchors:
            group_anchors[key] = r

    toc_row = header_last + 1
    # Сначала — unmerge в строке TOC (иначе ячейки в merged-диапазоне — read-only).
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == toc_row and mr.max_row == toc_row:
            ws.unmerge_cells(str(mr))
    # Затем чистим значения и гиперссылки старого оглавления.
    for c in range(1, TABLE_COLS + 1):
        cell = ws.cell(toc_row, c)
        cell.value = None
        cell.hyperlink = None

    write_toc(ws, toc_row, group_anchors)
    _save_with_backup(wb, svod_path, log=log)


def stage_set_heights_inplace(svod_path: Path, log=print) -> None:
    """Фиксирует высоты строк-заголовков в уже существующем своднике и
    включает wrap_text в столбцах H/N."""
    log(f"Фиксация высот + wrap: {svod_path.name}")
    wb = openpyxl.load_workbook(svod_path)
    if "Page1" not in wb.sheetnames:
        raise RuntimeError("В файле нет листа «Page1».")
    ws = wb["Page1"]
    header_last, data_last, _ = find_data_bounds(ws)
    label_set = {v.strip().lower() for v in GROUP_LABELS.values()}

    toc_row = header_last + 1
    if is_toc_row(ws, toc_row):
        ws.row_dimensions[toc_row].height = ROW_HEIGHT_TOC

    for r in range(header_last + 1, data_last + 1):
        if not is_section_row(ws, r):
            continue
        text = str(ws.cell(r, 1).value or "").strip().lower()
        if text in label_set:
            ws.row_dimensions[r].height = ROW_HEIGHT_SECTION
        else:
            ws.row_dimensions[r].height = ROW_HEIGHT_SUBSECTION
        _apply_section_vertical_center(ws, r)

    # wrap_text для H/N в строках данных + пересчёт высоты по тексту.
    for r in range(header_last + 1, data_last + 1):
        if is_toc_row(ws, r) or is_section_row(ws, r):
            continue
        if not is_equipment_row(ws, r):
            continue
        # Гарантируем стандартные объединения (на случай, если сводник был
        # собран ранней версией скрипта, где merge мог отсутствовать).
        ensure_equipment_merges(ws, r)
        for col in (8, 14):
            cell = ws.cell(r, col)
            al = cell.alignment
            if not al.wrap_text:
                cell.alignment = Alignment(
                    horizontal=al.horizontal, vertical=al.vertical,
                    text_rotation=al.text_rotation, wrap_text=True,
                    shrink_to_fit=al.shrink_to_fit, indent=al.indent,
                )
        a_text = str(ws.cell(r, 1).value or "")
        h_text = str(ws.cell(r, 8).value or "")
        n_text = str(ws.cell(r, 14).value or "")
        est = estimate_eq_row_height(ws, a_text, h_text, n_text)
        existing = ws.row_dimensions[r].height or 0.0
        ws.row_dimensions[r].height = max(existing, est)
        try:
            ws.row_dimensions[r].customHeight = True
        except Exception:
            pass

    _save_with_backup(wb, svod_path, log=log)


def stage_build_gantt_inplace(svod_path: Path, default_year: int | None = None,
                              log=print) -> None:
    """Пересоздаёт лист «Диаграмма» в уже существующем своднике.

    Читает актуальные даты начала/окончания (колонки F/G) прямо с Page1 —
    после ручного переноса сроков в Excel достаточно запустить эту стадию
    (GUI: «Обновить диаграмму Ганта» или чекбокс «Диаграмма Ганта»)."""
    log(f"Диаграмма Ганта: {svod_path.name}")
    wb = openpyxl.load_workbook(svod_path)
    if "Page1" not in wb.sheetnames:
        raise RuntimeError("В файле нет листа «Page1».")
    ws = wb["Page1"]

    parse_year = default_year_for_svod(svod_path, default_year)
    fn_month, fn_year = infer_schedule_from_filename(svod_path)
    recs = extract_records_from_svod(ws, default_year=parse_year)
    if not recs:
        raise RuntimeError(
            "На листе Page1 не найдено строк оборудования для диаграммы."
        )
    for rec in recs:
        g, sub = classify(rec)
        rec["group"] = g
        rec["subgroup"] = sub or rec.get("section", "")
    month, year = pick_month_year(recs, fn_year or default_year)
    if fn_month and not any(r.get("start") or r.get("end") for r in recs):
        month = fn_month

    log(f"  строк: {len(recs)}, шкала: {RU_MONTHS_NOM[month]} {year} г.")

    if GANTT_SHEET_NAME in wb.sheetnames:
        del wb[GANTT_SHEET_NAME]

    gantt_items = [{"row": r["src_row"], "group": r["group"], "rec": r}
                   for r in recs]
    build_gantt_sheet(wb, gantt_items, month, year)
    wb.active = wb.index(wb["Page1"])
    _save_with_backup(wb, svod_path, log=log)


# --- «Большие» стадии: полная пересборка ------------------------------------

def _load_inputs(root: Path, year_hint: int | None, log=print
                 ) -> tuple[dict, list[dict], Worksheet, int, int]:
    """Загружает справочник и проекты, возвращает (priority, records,
    template_ws, month, year). Падает с RuntimeError при проблемах."""
    p_prio = find_file(FILE_PRIO)
    p_arkh = find_file(FILE_ARKH)
    p_komi = find_file(FILE_KOMI)

    if not p_prio:
        raise RuntimeError(
            f"Не найден файл справочника «{FILE_PRIO}».\n"
            f"Положите его в папку: {root}"
        )
    if not p_arkh and not p_komi:
        raise RuntimeError(
            f"Не найдены ни «{FILE_ARKH}», ни «{FILE_KOMI}».\n"
            f"Положите хотя бы один из них в папку: {root}"
        )

    log("Найдены файлы:")
    log(f"  • {p_prio}")
    if p_arkh:
        log(f"  • {p_arkh}")
    if p_komi:
        log(f"  • {p_komi}")

    priority = load_priority(p_prio)
    default_year = year_hint if year_hint else datetime.now().year
    records: list[dict] = []
    ws_arkh = ws_komi = None

    if p_arkh:
        wb_arkh = openpyxl.load_workbook(p_arkh)
        ws_arkh = find_project_sheet(wb_arkh)
        layout_arkh = validate_project_template(ws_arkh, p_arkh.name)
        log(
            f"  формат «{p_arkh.name}»: лист «{layout_arkh.sheet_title}», "
            f"шапка до строки {layout_arkh.header_last}, "
            f"даты {get_column_letter(layout_arkh.col_start)}/"
            f"{get_column_letter(layout_arkh.col_end)}"
        )
        records += extract_records(
            ws_arkh, "Арх", default_year, "arkh", layout_arkh)

    if p_komi:
        wb_komi = openpyxl.load_workbook(p_komi)
        ws_komi = find_project_sheet(wb_komi)
        layout_komi = validate_project_template(ws_komi, p_komi.name)
        log(
            f"  формат «{p_komi.name}»: лист «{layout_komi.sheet_title}», "
            f"шапка до строки {layout_komi.header_last}, "
            f"даты {get_column_letter(layout_komi.col_start)}/"
            f"{get_column_letter(layout_komi.col_end)}"
        )
        records += extract_records(
            ws_komi, "Коми", default_year, "komi", layout_komi)

    log(f"Всего строк оборудования: {len(records)}")
    template_ws = ws_komi or ws_arkh
    month, year = pick_month_year(records, year_hint)
    log(f"Месяц сводного: {RU_MONTHS_NOM[month]} {year}")
    return priority, records, template_ws, month, year


def stage_full_rebuild(root: Path, year_hint: int | None,
                       opts: NormOptions, stats: NormStats,
                       log=print,
                       apply_sort: bool = True,
                       apply_toc: bool = True,
                       apply_heights: bool = True,
                       apply_gantt: bool = True) -> Path:
    """Собирает сводный график «с нуля» из проектов, со всеми выбранными
    стадиями. Возвращает путь к сохранённому файлу."""
    priority, records, template_ws, month, year = _load_inputs(
        root, year_hint, log=log)
    out_wb = build_output(
        priority, records, template_ws, None, month, year, opts, stats,
        apply_sort=apply_sort, apply_toc=apply_toc,
        apply_heights=apply_heights, apply_gantt=apply_gantt,
    )
    out_name = (
        f"{SVOD_FILE_PREFIX} ЛЭП и сетевого оборудования "
        f"на {RU_MONTHS_NOM[month]} {year} г.xlsx"
    )
    out_path = root / out_name
    _save_with_backup(out_wb, out_path, log=log)
    return out_path


def stage_rebuild_from_existing(svod_path: Path, year_hint: int | None,
                                opts: NormOptions, stats: NormStats,
                                log=print) -> Path:
    """Перечитывает существующий сводник и перестраивает его (полный набор
    стадий: расстановка приоритетов + TOC + высоты + Гант + нормализация,
    управляемая `opts`). Справочник приоритетов нужен обязательно.

    Стили и подписи берутся из самого сводника — он же и шаблон."""
    p_prio = find_file(FILE_PRIO)
    if not p_prio:
        raise RuntimeError(
            f"Не найден файл справочника «{FILE_PRIO}».\n"
            f"Положите его в папку: {ROOT}"
        )
    priority = load_priority(p_prio)

    log(f"Читаем существующий сводник: {svod_path.name}")
    wb = openpyxl.load_workbook(svod_path)
    if "Page1" not in wb.sheetnames:
        raise RuntimeError("В файле нет листа «Page1».")
    ws = wb["Page1"]

    default_year = default_year_for_svod(svod_path, year_hint)
    records = extract_records_from_svod(ws, default_year=default_year)
    log(f"Строк оборудования в своднике: {len(records)}")

    month, year = pick_month_year(records, year_hint)

    out_wb = build_output(
        priority, records, ws, None, month, year, opts, stats,
        apply_sort=True, apply_toc=True, apply_heights=True, apply_gantt=True,
    )
    # Имя файла оставляем тем же (месяц/год могут чуть измениться — тогда
    # возьмём новое имя). Бэкап старого выполнится в _save_with_backup.
    out_name = (
        f"{SVOD_FILE_PREFIX} ЛЭП и сетевого оборудования "
        f"на {RU_MONTHS_NOM[month]} {year} г.xlsx"
    )
    out_path = svod_path.parent / out_name
    # Если имя совпадает со старым — перезаписываем; если нет — старый тоже
    # бэкапим, чтобы не плодить разные копии.
    if out_path != svod_path and svod_path.exists():
        make_backup(svod_path, log=log)
    _save_with_backup(out_wb, out_path, log=log)
    return out_path


# ---------------------------------------------------------------------------
# Лист «Сравнение с проектами» — diff сводника vs исходные Коми/Арх
# ---------------------------------------------------------------------------

@dataclass
class DiffStats:
    """Сводка построения листа сравнения."""
    same: int = 0
    modified: int = 0
    new_in_svod: int = 0
    deleted_from_source: int = 0


def _norm_match_key(text: str) -> str:
    """Ключ сопоставления строк оборудования (без учёта регистра/пробелов)."""
    s = re.sub(r"\s+", " ", str(text or "").strip().lower())
    s = s.replace("–", "-").replace("—", "-")
    return s


def record_match_key(rec: dict) -> tuple[str, str]:
    return (str(rec.get("rdu") or ""), _norm_match_key(rec.get("name", "")))


def enrich_record_texts(rec: dict) -> dict:
    """Добавляет в запись текстовые поля A/H/N и даты для сравнения."""
    out = dict(rec)
    ws = rec["src_ws"]
    row = rec["src_row"]
    layout = rec.get("layout")
    if not isinstance(layout, ProjectLayout):
        layout = detect_project_layout(ws)
    out["layout"] = layout
    out["h_text"] = _cell_text_with_merges(ws, row, 8).strip()
    out["n_text"] = _cell_text_with_merges(ws, row, layout.col_repair).strip()
    out["start_text"] = _cell_text_with_merges(ws, row, layout.col_start).strip()
    out["end_text"] = _cell_text_with_merges(ws, row, layout.col_end).strip()
    return out


def load_source_records(root: Path, year_hint: int | None = None,
                        log=print) -> list[dict]:
    """Читает все строки оборудования из проектов Коми и Арх РДУ."""
    p_arkh = find_file(FILE_ARKH)
    p_komi = find_file(FILE_KOMI)
    if not p_arkh and not p_komi:
        raise RuntimeError(
            f"Для сравнения нужны «{FILE_ARKH}» и/или «{FILE_KOMI}» в папке {root}."
        )
    default_year = year_hint or datetime.now().year
    records: list[dict] = []
    if p_arkh:
        log(f"  исходник: {p_arkh.name}")
        wb = openpyxl.load_workbook(p_arkh)
        ws = find_project_sheet(wb)
        layout = validate_project_template(ws, p_arkh.name)
        records += extract_records(ws, "Арх", default_year, "arkh", layout)
    if p_komi:
        log(f"  исходник: {p_komi.name}")
        wb = openpyxl.load_workbook(p_komi)
        ws = find_project_sheet(wb)
        layout = validate_project_template(ws, p_komi.name)
        records += extract_records(ws, "Коми", default_year, "komi", layout)
    return [enrich_record_texts(r) for r in records]


def _dates_equal(a: tuple | None, b: tuple | None) -> bool:
    return a == b


def _record_fields_differ(svod: dict, source: dict) -> bool:
    if not _dates_equal(svod.get("start"), source.get("start")):
        return True
    if not _dates_equal(svod.get("end"), source.get("end")):
        return True
    if _norm_match_key(svod.get("h_text", "")) != _norm_match_key(source.get("h_text", "")):
        return True
    if _norm_match_key(svod.get("n_text", "")) != _norm_match_key(source.get("n_text", "")):
        return True
    if _norm_match_key(svod.get("name", "")) != _norm_match_key(source.get("name", "")):
        return True
    return False


def match_source_and_svod(source_recs: list[dict],
                          svod_recs: list[dict]
                          ) -> tuple[list[tuple[str, dict | None, dict | None]],
                                     DiffStats]:
    """Сопоставляет записи сводника с исходными проектами.

    Сначала по паре (РДУ, наименование), затем — только по наименованию
    среди оставшихся (на случай неточного определения РДУ в своднике).

    Возвращает список (status, svod_rec|None, source_rec|None) и статистику.
    status: same | modified | new | deleted
    """
    from collections import defaultdict

    pool: dict[tuple[str, str], list[dict]] = defaultdict(list)
    name_pool: dict[str, list[dict]] = defaultdict(list)
    for s in source_recs:
        pool[record_match_key(s)].append(s)
        name_pool[_norm_match_key(s.get("name", ""))].append(s)

    pairs: list[tuple[str, dict | None, dict | None]] = []
    stats = DiffStats()

    def _consume_source(src: dict) -> None:
        rk = record_match_key(src)
        nk = _norm_match_key(src.get("name", ""))
        for key, pd in ((rk, pool), (nk, name_pool)):
            lst = pd.get(key)
            if lst and src in lst:
                lst.remove(src)
                if not lst:
                    pd.pop(key, None)

    unmatched_sources: list[dict] = []
    for s in source_recs:
        unmatched_sources.append(s)

    for sv in svod_recs:
        src = None
        rk = record_match_key(sv)
        nk = _norm_match_key(sv.get("name", ""))
        lst = pool.get(rk)
        if lst:
            src = lst.pop(0)
            if not lst:
                pool.pop(rk, None)
        if src is None:
            lst = name_pool.get(nk)
            if lst:
                src = lst.pop(0)
                if not lst:
                    name_pool.pop(nk, None)
        if src is not None:
            _consume_source(src)
            if src in unmatched_sources:
                unmatched_sources.remove(src)
        if src is None:
            pairs.append(("new", sv, None))
            stats.new_in_svod += 1
        elif _record_fields_differ(sv, src):
            pairs.append(("modified", sv, src))
            stats.modified += 1
        else:
            pairs.append(("same", sv, src))
            stats.same += 1

    for src in unmatched_sources:
        pairs.append(("deleted", None, src))
        stats.deleted_from_source += 1

    return pairs, stats


def _inline_font(base: Font | None, *, color: str | None = None,
                 strike: bool = False, bold: bool = False) -> InlineFont:
    """InlineFont для rich text с наследованием размера/имени из ячейки."""
    name = (base.name if base and base.name else "Arial")
    size = base.size if base and base.size else 10.0
    kw: dict = {"rFont": name, "sz": size}
    if color:
        kw["color"] = color
    if strike:
        kw["strike"] = True
    if bold:
        kw["b"] = True
    return InlineFont(**kw)


def _text_diff_rich(old: str, new: str, base_font: Font | None) -> CellRichText | str:
    """Rich text: зелёные вставки, красный зачёркнутый удалённый фрагмент."""
    old = old or ""
    new = new or ""
    if old == new:
        return new
    sm = difflib.SequenceMatcher(None, old, new)
    blocks: list[TextBlock] = []
    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            chunk = new[j1:j2]
            if chunk:
                blocks.append(TextBlock(_inline_font(base_font), chunk))
        elif op == "delete":
            chunk = old[i1:i2]
            if chunk:
                blocks.append(TextBlock(
                    _inline_font(base_font, color=DIFF_COLOR_DEL, strike=True),
                    chunk,
                ))
        elif op == "insert":
            chunk = new[j1:j2]
            if chunk:
                blocks.append(TextBlock(
                    _inline_font(base_font, color=DIFF_COLOR_ADD),
                    chunk,
                ))
        elif op == "replace":
            ochunk = old[i1:i2]
            nchunk = new[j1:j2]
            if ochunk:
                blocks.append(TextBlock(
                    _inline_font(base_font, color=DIFF_COLOR_DEL, strike=True),
                    ochunk,
                ))
            if nchunk:
                blocks.append(TextBlock(
                    _inline_font(base_font, color=DIFF_COLOR_ADD),
                    nchunk,
                ))
    if not blocks:
        return new
    return CellRichText(blocks)


def _format_date_change(old_t: str, new_t: str, old_d: tuple | None,
                        new_d: tuple | None, base_font: Font | None
                        ) -> CellRichText | str:
    """Ячейка даты: новое значение + подсветка; при изменении — было → стало."""
    new_show = new_t or (format_date_tuple(new_d) if new_d else "")
    old_show = old_t or (format_date_tuple(old_d) if old_d else "")
    if _norm_match_key(old_show) == _norm_match_key(new_show):
        return new_show
    if not old_show:
        return _text_diff_rich("", new_show, base_font)
    if not new_show:
        return CellRichText([
            TextBlock(_inline_font(base_font, color=DIFF_COLOR_DEL, strike=True),
                      old_show),
        ])
    combined_old = f"{old_show} → "
    combined_new = f"{old_show} → {new_show}"
    return _text_diff_rich(combined_old, combined_new, base_font)


def format_date_tuple(d: tuple[int, int, int] | None) -> str:
    if not d:
        return ""
    _y, m, day = d
    return f"{day:02d}.{m:02d}."


def _apply_row_fill(ws: Worksheet, row: int, color: str,
                    ncols: int = TABLE_COLS) -> None:
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in range(1, ncols + 1):
        ws.cell(row, c).fill = fill


def _apply_strikethrough_row(ws: Worksheet, row: int,
                             ncols: int = TABLE_COLS) -> None:
    """Красный зачёркнутый текст во всех непустых ячейках строки."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        v = cell.value
        if v is None or str(v).strip() == "":
            continue
        base = cell.font
        cell.font = Font(
            name=base.name if base else "Arial",
            size=base.size if base else 10,
            bold=base.bold if base else False,
            italic=base.italic if base else False,
            strike=True,
            color=DIFF_COLOR_DEL,
        )


def _write_diff_legend(ws: Worksheet) -> None:
    """Легенда в правом верхнем углу листа сравнения."""
    legend = (
        "Легенда:  "
        "только изменённые строки;  "
        "зелёный — добавленный текст;  "
        "красный зачёркнутый — удалённый;  "
        "жёлтая заливка — изменённые даты;  "
        "зелёная строка — новая в своднике;  "
        "красная строка — удалена из проекта"
    )
    cell = ws.cell(1, TABLE_COLS + 1)
    cell.value = legend
    cell.font = Font(name="Arial", size=9, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[get_column_letter(TABLE_COLS + 1)].width = 48


def _annotate_equipment_row_diff(ws: Worksheet, row: int, svod: dict,
                                 source: dict | None, status: str) -> None:
    """Размечает одну строку оборудования на листе сравнения."""
    layout = svod.get("layout") or detect_project_layout(ws)

    if status == "new":
        _apply_row_fill(ws, row, DIFF_FILL_NEW_ROW)
        return

    if status == "same" or source is None:
        return

    # Текстовые колонки A, H, N — посимвольный diff.
    for col, fld in ((1, "name"), (8, "h_text"), (layout.col_repair, "n_text")):
        cell = ws.cell(row, col)
        old_t = source.get(fld, "")
        new_t = svod.get(fld, "")
        if _norm_match_key(old_t) != _norm_match_key(new_t):
            rich = _text_diff_rich(old_t, new_t, cell.font)
            cell.value = rich

    # Даты F/G.
    for col, fld_t, fld_d in (
        (layout.col_start, "start_text", "start"),
        (layout.col_end, "end_text", "end"),
    ):
        cell = ws.cell(row, col)
        old_t = source.get(fld_t, "")
        new_t = svod.get(fld_t, "")
        if (not _dates_equal(svod.get(fld_d), source.get(fld_d))
                or _norm_match_key(old_t) != _norm_match_key(new_t)):
            cell.value = _format_date_change(
                old_t, new_t, source.get(fld_d), svod.get(fld_d), cell.font,
            )
            cell.fill = PatternFill(
                start_color=DIFF_FILL_DATE_CHG,
                end_color=DIFF_FILL_DATE_CHG,
                fill_type="solid",
            )


def _insert_deleted_source_rows(ws: Worksheet, deleted: list[dict],
                                insert_before: int, style_row: int | None,
                                log=print) -> int:
    """Добавляет в конец diff-листа строки, удалённые из сводника."""
    if not deleted:
        return 0
    cur = insert_before
    hdr_font = Font(name="Arial", size=10, bold=True, color=DIFF_COLOR_DEL)
    hdr = ws.cell(cur, 1)
    hdr.value = (
        f"─── Удалено из исходных проектов ({len(deleted)} стр., "
        f"не вошло в сводник) ───"
    )
    hdr.font = hdr_font
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    _apply_row_fill(ws, cur, DIFF_FILL_DELETED_ROW)
    try:
        ws.merge_cells(f"A{cur}:{LAST_COL_LETTER}{cur}")
    except Exception:
        pass
    cur += 1
    for src in deleted:
        copy_row_full(src["src_ws"], src["src_row"], ws, cur)
        copy_merges_in_row(src["src_ws"], src["src_row"], ws, cur)
        ensure_equipment_merges(ws, cur)
        _apply_row_fill(ws, cur, DIFF_FILL_DELETED_ROW)
        _apply_strikethrough_row(ws, cur)
        tag = ws.cell(cur, TABLE_COLS + 1)
        tag.value = f"[удалено · {src.get('rdu', '?')} РДУ]"
        tag.font = Font(name="Arial", size=8, color=DIFF_COLOR_DEL, italic=True)
        cur += 1

    log(f"  + {len(deleted)} удалённых строк из проектов")
    return cur - insert_before


def _copy_header_block(src_ws: Worksheet, dst_ws: Worksheet,
                       header_last: int) -> None:
    """Копирует шапку (строки 1..header_last) с объединениями."""
    for r in range(1, header_last + 1):
        copy_row_full(src_ws, r, dst_ws, r)
    for mr in src_ws.merged_cells.ranges:
        if mr.min_row <= header_last and mr.max_row <= header_last:
            rng = (
                f"{get_column_letter(mr.min_col)}{mr.min_row}:"
                f"{get_column_letter(min(mr.max_col, TABLE_COLS))}{mr.max_row}"
            )
            try:
                dst_ws.merge_cells(rng)
            except Exception:
                pass


def _copy_data_row_with_merges(src_ws: Worksheet, dst_ws: Worksheet,
                               src_row: int, dst_row: int) -> None:
    """Копирует одну строку данных со стилями и однострочными merge."""
    copy_row_full(src_ws, src_row, dst_ws, dst_row)
    copy_merges_in_row(src_ws, src_row, dst_ws, dst_row)
    if is_equipment_row(src_ws, src_row):
        ensure_equipment_merges(dst_ws, dst_row)


def _write_diff_filter_note(ws: Worksheet, row: int) -> None:
    """Пояснение под шапкой: на листе только изменённые строки."""
    text = (
        "На этом листе показаны только изменённые строки "
        "(без неизменённых). См. легенду справа."
    )
    cell = ws.cell(row, 1)
    cell.value = text
    cell.font = Font(name="Arial", size=9, italic=True, color="444444")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    try:
        ws.merge_cells(f"A{row}:{LAST_COL_LETTER}{row}")
    except Exception:
        pass
    ws.row_dimensions[row].height = 16.0


def _write_diff_empty_note(ws: Worksheet, row: int) -> None:
    cell = ws.cell(row, 1)
    cell.value = "Изменений относительно исходных проектов Коми/Арх РДУ не найдено."
    cell.font = Font(name="Arial", size=10, italic=True)
    try:
        ws.merge_cells(f"A{row}:{LAST_COL_LETTER}{row}")
    except Exception:
        pass


def build_diff_sheet(wb: openpyxl.Workbook, svod_ws: Worksheet,
                     source_recs: list[dict], svod_recs: list[dict],
                     log=print) -> DiffStats:
    """Создаёт лист «Сравнение с проектами» — только изменённые строки."""
    if DIFF_SHEET_NAME in wb.sheetnames:
        del wb[DIFF_SHEET_NAME]
    diff_ws = wb.create_sheet(DIFF_SHEET_NAME)
    copy_column_widths(svod_ws, diff_ws)

    header_last, _data_last, sig_start = find_data_bounds(svod_ws)
    _copy_header_block(svod_ws, diff_ws, header_last)
    dst_row = header_last + 1
    _write_diff_filter_note(diff_ws, dst_row)
    dst_row += 1

    svod_enriched = [enrich_record_texts(r) for r in svod_recs]
    pairs, stats = match_source_and_svod(source_recs, svod_enriched)

    pair_by_row: dict[int, tuple[str, dict, dict | None]] = {}
    deleted: list[dict] = []
    for status, sv, src in pairs:
        if status == "deleted" and src:
            deleted.append(src)
        elif sv is not None:
            pair_by_row[sv["src_row"]] = (status, sv, src)

    changed_count = stats.modified + stats.new_in_svod
    if changed_count == 0 and not deleted:
        _write_diff_empty_note(diff_ws, dst_row)
        _write_diff_legend(diff_ws)
        log(f"  изменений нет ({stats.same} строк без отличий)")
        return stats

    pending_sections: list[int] = []
    copied_sections: set[int] = set()

    for src_row in range(header_last + 1, sig_start):
        if is_toc_row(svod_ws, src_row):
            continue
        if is_section_row(svod_ws, src_row):
            pending_sections.append(src_row)
            continue
        if not is_equipment_row(svod_ws, src_row):
            continue

        entry = pair_by_row.get(src_row)
        if not entry or entry[0] == "same":
            continue

        status, sv, src = entry
        for sec_r in pending_sections:
            if sec_r not in copied_sections:
                _copy_data_row_with_merges(svod_ws, diff_ws, sec_r, dst_row)
                copied_sections.add(sec_r)
                dst_row += 1
        pending_sections = []

        _copy_data_row_with_merges(svod_ws, diff_ws, src_row, dst_row)
        _annotate_equipment_row_diff(diff_ws, dst_row, sv, src, status)
        dst_row += 1

    if deleted:
        _insert_deleted_source_rows(diff_ws, deleted, dst_row, None, log=log)

    _write_diff_legend(diff_ws)
    log(
        f"  на листе: {changed_count} изменённых/новых + "
        f"{stats.deleted_from_source} удалённых "
        f"(скрыто без изменений: {stats.same})"
    )
    return stats


def stage_build_diff_inplace(svod_path: Path, root: Path | None = None,
                             year_hint: int | None = None,
                             log=print) -> DiffStats:
    """Добавляет в сводник лист «Сравнение с проектами».

    **За основу берётся лист Page1 указанного сводника** (обычно файл
    «Сводный график …xlsx» из корня папки программы — с учётом ручных правок).
    Сравнение идёт с «Проект Коми РДУ.xlsx» и «Проект Арх РДУ.xlsx» из корня
    (или из «Исходные материалы/», если в корне их нет).

    На листе сравнения — **только изменённые строки** (без неизменённых):
      • зелёный текст — добавленные символы;
      • красный зачёркнутый — удалённый фрагмент / строка;
      • жёлтая заливка — изменённые даты начала/окончания.
    """
    root = root or ROOT
    log(f"Сравнение с проектами: {svod_path.name}")
    log(f"  за основу: {svod_path} (лист Page1)")
    log("  эталон: проекты Коми/Арх РДУ из папки программы")
    parse_year = default_year_for_svod(svod_path, year_hint)
    source_recs = load_source_records(root, parse_year, log=log)

    wb = openpyxl.load_workbook(svod_path)
    if "Page1" not in wb.sheetnames:
        raise RuntimeError("В файле нет листа «Page1».")
    ws = wb["Page1"]
    svod_recs = extract_records_from_svod(ws, default_year=parse_year)

    stats = build_diff_sheet(wb, ws, source_recs, svod_recs, log=log)
    log(
        f"  без изменений: {stats.same}; изменено: {stats.modified}; "
        f"новых в своднике: {stats.new_in_svod}; "
        f"удалено из проектов: {stats.deleted_from_source}"
    )
    log(f"  лист «{DIFF_SHEET_NAME}» создан.")
    _save_with_backup(wb, svod_path, log=log)
    return stats


# ---------------------------------------------------------------- ТОЧКА ВХОДА

def _short(s: str, limit: int = 100) -> str:
    """Укорачивает многострочный текст до одной строки ≤ limit символов."""
    if s is None:
        return ""
    s = str(s).replace("\n", " ⏎ ")
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > limit:
        s = s[: limit - 1] + "…"
    return s


def _print_norm_report(stats: NormStats, dry_run: bool) -> None:
    """Печатает отчёт о применённых правилах нормализации текста."""
    print()
    print("Нормализация текста:")
    if not stats.counts:
        print("  изменений нет.")
    else:
        for label, c in sorted(stats.counts.items(), key=lambda kv: (-kv[1], kv[0])):
            print(f"  • {label}: {c}")

    if dry_run:
        print()
        print("Детализация изменений (--dry-run, файл не сохранён):")
        if not stats.changes:
            print("  нет.")
        for ch in stats.changes:
            print(f"  {ch['row_label']}")
            if ch['h_before'] != ch['h_after']:
                print(f"    H: {_short(ch['h_before'])}")
                print(f"     →  {_short(ch['h_after']) or '(пусто)'}")
            if ch['n_before'] != ch['n_after']:
                print(f"    N: {_short(ch['n_before'])}")
                print(f"     →  {_short(ch['n_after']) or '(пусто)'}")


STAGE_CHOICES = (
    "all",         # полная пересборка с нуля (по умолчанию)
    "merge",       # только объединение проектов (без сортировки/TOC/Ганта/высот)
    "sort",        # перечитать существующий сводник и переставить по приоритетам
    "normalize",   # только нормализация текста в готовом своднике
    "toc",         # только перегенерация оглавления
    "heights",     # только фиксация высот и wrap_text
    "gantt",       # только перестроить лист «Диаграмма»
    "diff",        # лист «Сравнение с проектами» vs исходные Коми/Арх
    "restore",     # откатить сводник к последней резервной копии
)


def _require_existing_svod(log=print) -> Path:
    """Возвращает путь к существующему своднику в корне или падает с понятной
    ошибкой."""
    path = find_existing_svod(ROOT)
    if path is None:
        raise RuntimeError(
            f"В папке {ROOT} не найден файл «{SVOD_FILE_PREFIX} …xlsx».\n"
            f"Сначала выполните стадию «merge» или «all»."
        )
    return path


def main():
    parser = argparse.ArgumentParser(
        description="Сборщик сводного графика ремонтов ЛЭП и сетевого оборудования.",
        epilog=__copyright__,
    )
    parser.add_argument("--stage", choices=STAGE_CHOICES, default="all",
                        help="Какую стадию выполнить. По умолчанию «all» — "
                             "полная пересборка из проектов.")
    parser.add_argument("--year", type=int, default=None,
                        help="Год в имени выходного файла (по умолчанию — из дат проекта или текущий).")
    parser.add_argument("--no-normalize", action="store_true",
                        help="Отключить текстовую нормализацию полей H и N "
                             "(применяется к «all», «merge», «sort»).")
    parser.add_argument("--collapse-preamble", action="store_true",
                        help="Сворачивать преамбулы «Вывод в ремонт … для проведения …» "
                             "в краткую форму «<Вид ремонта> Y» (опытное правило).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Ничего не сохранять — только показать, что будет изменено "
                             "(работает для «all» и «merge»).")
    args = parser.parse_args()

    opts = NormOptions(
        enabled=not args.no_normalize,
        collapse_preamble=bool(args.collapse_preamble),
        dry_run=bool(args.dry_run),
    )
    stats = NormStats()

    try:
        if args.stage == "all":
            if opts.dry_run:
                # «Сухой прогон» собираем в памяти, но не сохраняем.
                priority, records, tws, month, year = _load_inputs(
                    ROOT, args.year, log=print)
                build_output(priority, records, tws, None, month, year,
                             opts, stats)
                _print_norm_report(stats, dry_run=True)
                print()
                print("[--dry-run] Итоговый файл не сохранён.")
                return
            out_path = stage_full_rebuild(
                ROOT, args.year, opts, stats,
                apply_sort=True, apply_toc=True,
                apply_heights=True, apply_gantt=True,
            )
            _print_norm_report(stats, dry_run=False)
            print(f"\nГотово: {out_path}")

        elif args.stage == "merge":
            # Чистое объединение: классификация без приоритетов, без TOC/высот/Ганта.
            out_path = stage_full_rebuild(
                ROOT, args.year, opts, stats,
                apply_sort=False, apply_toc=False,
                apply_heights=False, apply_gantt=False,
            )
            _print_norm_report(stats, dry_run=False)
            print(f"\nГотово: {out_path}")

        elif args.stage == "sort":
            svod = _require_existing_svod()
            out_path = stage_rebuild_from_existing(
                svod, args.year, opts, stats, log=print)
            _print_norm_report(stats, dry_run=False)
            print(f"\nГотово: {out_path}")

        elif args.stage == "normalize":
            svod = _require_existing_svod()
            stage_normalize_inplace(svod, opts, stats, log=print)
            _print_norm_report(stats, dry_run=False)

        elif args.stage == "toc":
            svod = _require_existing_svod()
            stage_build_toc_inplace(svod, log=print)

        elif args.stage == "heights":
            svod = _require_existing_svod()
            stage_set_heights_inplace(svod, log=print)

        elif args.stage == "gantt":
            svod = _require_existing_svod()
            stage_build_gantt_inplace(svod, args.year, log=print)

        elif args.stage == "diff":
            svod = _require_existing_svod()
            stage_build_diff_inplace(svod, ROOT, args.year, log=print)

        elif args.stage == "restore":
            svod = find_existing_svod(ROOT)
            if svod is None:
                # Нечего откатывать в корне — попробуем восстановить по любой
                # копии: возьмём ту, к чьему имени больше всего копий.
                raise RuntimeError(
                    f"В папке {ROOT} не найден свод. Сначала положите файл "
                    f"«{SVOD_FILE_PREFIX} …xlsx» или запустите стадию «merge»/"
                    f"«all»."
                )
            restored = restore_latest_backup(svod, log=print)
            if restored is None:
                sys.exit(4)

    except RuntimeError as e:
        print(f"ОШИБКА: {e}")
        sys.exit(2)


if __name__ == "__main__":
    main()
